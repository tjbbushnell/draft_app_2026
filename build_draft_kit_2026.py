#!/usr/bin/env python3
"""
=============================================================================
 2026 FANTASY DRAFT KIT BUILDER  --  v8 (bye week column)
=============================================================================
Builds "Draft_Kit_2026.xlsx", hardcoded to one specific league:

    10-team, 2026 Season Redraft, Sleeper, Snake draft, pick 8 of 16 rounds.
    Roster: 1 QB / 2 RB / 2 WR / 1 TE / 2 FLEX (W/R/T) / 1 SUPERFLEX (W/R/T/Q)
            / 7 BN / 2 IR
    Scoring: 0.5 PPR, 1.0 TE Premium, custom passing/rushing/receiving
             milestone bonuses and penalties (full table on the "Scoring
             Rules" sheet; a live "Scoring Calculator" sheet turns any raw
             stat line into this league's exact point total).

WHAT'S NEW IN THIS VERSION (vs. the v4 global-VBD-board draft kit)
-----------------------------------------------------------------------
1. REAL ADP + VALUE DELTA. A new ADP column carries real 2026 Superflex
   Half-PPR Average Draft Position (FantasyPros Real-Time ADP, 250 real
   players matched in by name -- see section 3b for provenance/caveats). A
   new VALUE DELTA column = ADP - GlobalRank: a big POSITIVE number means
   the market is letting that player slide past where this model has him
   (a real draft-day discount), a big NEGATIVE number flags a likely reach.
   Both colors code automatically on the sheet.
2. THE DRAFT BOARD IS NOW A REAL, NATIVE, FULLY FILTERABLE EXCEL TABLE.
   The old layout (v2-v4) interleaved merged tier-banner rows among the
   player rows -- that looks nice but Excel CANNOT natively sort or filter
   a range containing merged cells without breaking it. So the tier
   banners, live counts, and Draft Pivot & Action Rule text all moved to a
   new, ALWAYS-VISIBLE "TIER GUIDE" block at the very top of the sheet
   (still fully live -- its counts still update as you check off players),
   and the actual 328-player list below it is now a uniform, single-
   header-row Excel Table (native filter dropdowns on every column, safe
   to sort by ADP/POS/PLAYER/anything, banded rows). A new persistent
   TIER LEVEL text column ("Tier 1", "Tier 2", ...) rides on every player
   row specifically so a player's tier stays visible no matter how you've
   sorted or filtered the table.
3. Everything else carries forward unchanged and re-verified: the real
   FantasyPros ECR baseline, the global VBD interleaving and tiering, real
   Weeks-1-4-only SoS, real checkboxes + live strikethrough formatting,
   Contract-Year flags, the Superflex/TEP tag column, and the 4 core tabs
   (Draft Board, Scoring Rules incl. the Global Value model at the bottom,
   Scoring Calculator, Instructions).

WHAT'S NEW IN v6 (the morning-of-draft refresh, 2026-08-30)
-----------------------------------------------------------------------
No structural changes -- same 17 columns, same 4 tabs, same VBD/ADP model.
This pass re-researched every player who came up in the mock-draft review
(fresh camp reports, injury news, trades, and role battles as of 8/29-8/30)
and corrected/added their Flag + Notes: Bucky Irving and Matthew Golden were
too pessimistic (both have real positive camp news now); David Montgomery,
Rhamondre Stevenson, Kenny Gainwell, Breece Hall, Chris Olave, Jonah Coleman,
Jalen Coker, Tyler Allgeier, and Dallas Goedert got corrected/added
committee-risk or injury notes; Bryce Young, Brian Thomas Jr., Jakobi
Meyers, Michael Pittman Jr., Rashid Shaheed, Jaylen Waddle, and Romeo Doubs
(all previously ECR-only, no situational note) got one added. Also added a
"DRAFT-MORNING WATCH LIST" section at the top of the Instructions tab and a
callout about the ADP column's real TE-premium blind spot (see below).

WHAT'S NEW IN v7 (the same night -- a scoped, deliberately-conservative
upgrade to the value model itself, after a critical review of a much more
ambitious rewrite proposal)
-----------------------------------------------------------------------
A fuller rewrite was proposed (drop ECR entirely for per-player statistical
projections, fit the model's constants to ADP via an optimizer, make the
whole board dynamically recalculate as players are drafted, add a full-
roster injury-durability discount). The honest review of that proposal is
in this project's chat history; the short version is that two of its four
ideas were legitimate but not safe to build and ship untested hours before
a real draft (fabricating per-player projections we don't have real data
for, and a live-recalculating engine substituted in at the last minute for
a static workbook). We shipped the safe half instead:
  1. SoS's effect on GLOBAL VALUE is now a CONTINUOUS function of SoS_Rank
     (1-32) instead of a flat bonus keyed to a discrete 1-5 tier -- see
     _sos_adj_continuous(). No new data required, this data already existed.
  2. A narrow, hand-curated DURABILITY discount (see the DURABILITY dict)
     applied ONLY to the ~16 players whose Notes already cite a real,
     specific injury/durability concern -- NOT a full-roster feature,
     deliberately, since we have no real games-missed data for the other
     ~310 players and won't fabricate it.
  3. A "ROSTER-AWARE MARGINAL VALUE GUIDE" added to the Instructions tab --
     a precomputed, static reference (not a live recalculation) for
     discounting a player's GLOBAL VALUE by how many you already have at
     that position, so the "don't draft a 3rd TE" lesson from the mock-draft
     review is something you can apply by eye without any new engineering.
The ADP-calibration idea (fit PEAK_POINTS/DECAY_ALPHA to real market ADP)
was NOT implemented tonight either -- it has a real internal conflict with
preserving the TE premium divergence that needs more careful design than
one evening allows. GLOBAL VALUE numbers will shift slightly from v6
because of the SoS and durability changes; GlobalRank/Tier/TierLevel/
ValueDelta all flow from GlobalValue so they may re-sort slightly too.

WHAT'S NEW IN v8
-----------------------------------------------------------------------
Added a BYE column (real 2026 NFL bye week per team, verified against two
independent sources -- NFL.com's own schedule-release article and Sports
Illustrated, cross-checked for an exact match; see BYE_WEEKS). Sits right
after TEAM on the Draft Board table. Purely informational -- it does not
feed into GLOBAL VALUE or tiering at all, since bye-week roster construction
(making sure your starters don't all share a bye) is a decision you make
across your whole roster during/after the draft, not a single-player value
adjustment. No other changes.

WHAT'S NEW IN v9
-----------------------------------------------------------------------
Two draft-morning requests. (1) BYE column confirmed to cover all 328
players: 324 real players get their team's actual bye week; the 4 rows
with Team=="FA" (Tyreek Hill, Kareem Hunt, Joe Mixon, Cedric Tillman) now
show "FA" instead of a blank cell, since a free agent genuinely has no bye
week to report -- this is a correct, verified absence, not missing data.
(2) Relocated the TIER GUIDE (composition / cliffs / live remaining counts /
Draft Pivot & Action Rule per tier) off the Draft Board sheet and onto the
top of the Instructions tab, keeping the workbook at exactly 4 tabs (per
explicit preference -- a 5th "Tier Guide" tab was offered and declined).
The Draft Board sheet is now just a one-line pointer banner (row 1) plus
the clean, fully sortable/filterable player table -- nothing else. The
Tier Guide's live remaining-count formulas now use cross-sheet COUNTIFS
references (e.g. 'Draft Board'!C4:C331) since the guide and the table it
counts now live on different sheets; verified via LibreOffice recalculation
that checking off a player on Draft Board still live-updates the correct
tier's remaining count on Instructions. No changes to the value model,
scoring, SoS, or durability logic in this version.

HOW TO USE
----------
    pip install pandas openpyxl
    python3 build_draft_kit_2026.py

Outputs, in the same folder:
    players_2026.csv     -- full 328-player pool, source of truth
    Draft_Kit_2026.xlsx  -- the finished interactive workbook

DATA PROVENANCE / KNOWN LIMITS
-----------------------------------------------------------------------
- ECR + team: FantasyPros Half-PPR Overall Draft ECR, 108 experts, fetched
  2026-08-29 (this is STANDARD half-PPR ECR -- FantasyPros has no
  TE-Premium-specific scoring toggle, confirmed by direct search; the TEP
  reweighting is handled separately via the Adj. Value Score multiplier,
  not by a different ECR source).
- A Superflex-specific overall ECR exists on FantasyPros but is login-walled
  after rank 20; those 20 real superflex ranks are captured and shown as a
  bonus "SF #n" tag, they do not override the primary ECR-based sort.
- Situational tags (CY / Flag / Notes) are only as fresh as the week-of
  Aug 24-29, 2026 research pass -- re-verify anything time-sensitive
  (injuries, suspensions, depth charts) the morning of your draft.
- Early-Season SoS: the Weeks 1-4 schedule itself is confirmed/cross-checked
  (Pro-Football-Reference + FootballDB agree on every 2026 Wk1-4 matchup).
  The defense-strength inputs behind it are the honest limitation: no free,
  forward-looking 2026 pass-D/run-D SPLIT projection exists publicly (PFF
  and FTN DVOA both paywall that exact split). The script uses each team's
  actual final-2025 pass/rush yards allowed (Pro-Football-Reference) as the
  strength proxy, cross-checked directionally against Action Network's
  Aug-28-2026 qualitative preseason defense tier rankings (broad agreement
  on the best/worst tiers). Treat "Early SoS" as a real, sourced, but
  proxy-based signal -- not a paid 2026 DVOA model -- and re-verify against
  live odds/projections closer to kickoff if a specific matchup matters a
  lot to a roster decision.
- GLOBAL VALUE MODEL: converts each player's real ECR rank into an
  estimated full-season point total using a position-specific power-law
  decay curve (PEAK_POINTS, DECAY_ALPHA), then subtracts a replacement-
  level baseline (REPLACEMENT_RANK, an explicit assumption about how many
  players per position actually start in this league's 9-starter/2-FLEX/
  1-SUPERFLEX roster -- documented next to the constants). The result
  (Value-Based Drafting, or VBD) is what makes QB/RB/WR/TE comparable on
  one scale, and is the ONLY thing that should be trusted for "who's the
  better player regardless of position" -- ECR alone never answers that,
  since it's position-relative by definition. The PEAK/ALPHA/REPLACEMENT
  constants are calibrated to realistic, well-known half-PPR/TE-premium
  scoring shapes (elite RB/TE cliffs hard, QB/WR stay deep longer, TE
  premium closes most but not all of the gap to elite WR value) -- they
  are estimates, not official 2026 point projections, and are yours to
  edit (see the constants block and the "Scoring Rules" tab) if your own
  read on positional value differs.
- ADP: FantasyPros Real-Time ADP, Superflex view, Half-PPR, fetched
  2026-08-29 ("last updated 5 minutes ago" at fetch time) -- confirmed
  genuinely Superflex-formatted (QBs pulled way up vs. standard ADP: Josh
  Allen 4.8 overall, Lamar Jackson 8.9). Two honest caveats: it's a
  12-team ADP tool (no 10-team toggle was available), so treat the ADP
  column as a relative ranking more than a literal 10-team pick slot; and
  it caps at 250 rows, so the deepest ~78 players in the 328-player pool
  have no ADP at all (expected -- they're below what live rooms are
  actually drafting). VALUE DELTA (ADP - GlobalRank) is therefore blank
  for those players too.
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

CSV_PATH = "players_2026.csv"
XLSX_PATH = "Draft_Kit_2026.xlsx"
N_TIERS = 6          # per-position "POS TIER" reference column (unchanged from before)
N_GLOBAL_TIERS = 12  # primary Global Tier column -- one board, 328 players, finer granularity

# =============================================================================
# 1. LEAGUE / SCORING RULES
# =============================================================================

LEAGUE_INFO = [
    ("Platform", "Sleeper (Enhanced Sleeper Sync), 10-team Redraft, 2026 season"),
    ("Draft format", "Snake draft, 16 rounds, you draft 8th overall"),
    ("Starting roster", "1 QB, 2 RB, 2 WR, 1 TE, 2 FLEX (W/R/T), 1 SUPERFLEX (W/R/T/Q)"),
    ("Bench / IR", "7 Bench, 2 IR"),
    ("PPR base", "0.5 PPR (all positions)"),
    ("TE Premium", "TEs get an ADDITIONAL +0.5 / reception on top of the 0.5 base = 1.0 total PPR for TEs only"),
]

SCORING_RULES = [
    ("PASSING", "Yards", "+0.04 / yard (25 yards = 1 pt)"),
    ("PASSING", "Touchdowns", "+5.0 / passing TD"),
    ("PASSING", "40+ yard TD bonus", "+0.5 additional, on top of the +5"),
    ("PASSING", "Interceptions", "-2.0 / INT (this is the TOTAL INT count, pick-6es included)"),
    ("PASSING", "Pick-6 thrown", "-1.0 additional per pick-6, on top of the -2 INT penalty"),
    ("PASSING", "Sacks taken", "-0.5 / sack"),
    ("RUSHING", "Yards", "+0.1 / yard (10 yards = 1 pt)"),
    ("RUSHING", "Touchdowns", "+6.0 / rushing TD"),
    ("RUSHING", "40+ yard TD bonus", "+2.0 additional, on top of the +6"),
    ("RUSHING", "2-pt conversion", "+2.0"),
    ("RUSHING", "100-199 yard game bonus", "+2.0 (milestone, does not stack with the 200+ bonus)"),
    ("RUSHING", "200+ yard game bonus", "+4.0 (milestone, replaces the 100-199 bonus, doesn't stack)"),
    ("RECEIVING", "Reception (base)", "+0.5 / catch (all positions)"),
    ("RECEIVING", "Reception (TE Premium)", "+1.0 / catch total for TEs only (0.5 base + 0.5 TEP)"),
    ("RECEIVING", "Yards", "+0.1 / yard (10 yards = 1 pt)"),
    ("RECEIVING", "Touchdowns", "+6.0 / receiving TD"),
    ("RECEIVING", "40+ yard TD bonus", "+2.0 additional, on top of the +6"),
    ("RECEIVING", "100-199 yard game bonus", "+3.0 (milestone, does not stack with the 200+ bonus)"),
    ("RECEIVING", "200+ yard game bonus", "+5.0 (milestone, replaces the 100-199 bonus, doesn't stack)"),
    ("MISC", "Fumble", "-1.0 (total fumbles)"),
    ("MISC", "Fumble lost", "-1.0 additional per fumble lost, on top of the -1 fumble penalty"),
]

# =============================================================================
# 2. STRENGTH OF SCHEDULE -- REAL, ISOLATED WEEKS 1-4 (not a season-long
#    blend). Two real inputs, combined below into a genuine early-season
#    metric. Team codes normalized to match the FantasyPros ECR pull: JAC, LV.
# =============================================================================

# 2a. Real 2026 Weeks 1-4 opponents for all 32 teams (Pro-Football-Reference
#     + FootballDB, cross-verified: every week's 16 matchups reconcile
#     perfectly both ways).
SCHEDULE_WK1_4 = {
    "ARI": ["LAC", "SEA", "SF", "NYG"], "ATL": ["PIT", "CAR", "GB", "NO"],
    "BAL": ["IND", "NO", "DAL", "TEN"], "BUF": ["HOU", "DET", "LAC", "NE"],
    "CAR": ["CHI", "ATL", "CLE", "DET"], "CHI": ["CAR", "MIN", "PHI", "NYJ"],
    "CIN": ["TB", "HOU", "PIT", "JAC"], "CLE": ["JAC", "TB", "CAR", "PIT"],
    "DAL": ["NYG", "WAS", "BAL", "HOU"], "DEN": ["KC", "JAC", "LAR", "SF"],
    "DET": ["NO", "BUF", "NYJ", "CAR"], "GB": ["MIN", "NYJ", "ATL", "TB"],
    "HOU": ["BUF", "CIN", "IND", "DAL"], "IND": ["BAL", "KC", "HOU", "WAS"],
    "JAC": ["CLE", "DEN", "NE", "CIN"], "KC": ["DEN", "IND", "MIA", "LV"],
    "LAC": ["ARI", "LV", "BUF", "SEA"], "LAR": ["SF", "NYG", "DEN", "PHI"],
    "LV": ["MIA", "LAC", "NO", "KC"], "MIA": ["LV", "SF", "KC", "MIN"],
    "MIN": ["GB", "CHI", "TB", "MIA"], "NE": ["SEA", "PIT", "JAC", "BUF"],
    "NO": ["DET", "BAL", "LV", "ATL"], "NYG": ["DAL", "LAR", "TEN", "ARI"],
    "NYJ": ["TEN", "GB", "DET", "CHI"], "PHI": ["WAS", "TEN", "CHI", "LAR"],
    "PIT": ["ATL", "NE", "CIN", "CLE"], "SEA": ["NE", "ARI", "WAS", "LAC"],
    "SF": ["LAR", "MIA", "ARI", "DEN"], "TB": ["CIN", "CLE", "MIN", "GB"],
    "TEN": ["NYJ", "PHI", "NYG", "BAL"], "WAS": ["PHI", "DAL", "SEA", "IND"],
}

# 2b. Real per-team defensive strength, 1 = toughest .. 32 = weakest, split
#     pass vs. run. Proxy = actual final-2025 yards allowed (Pro-Football-
#     Reference), cross-checked directionally against Action Network's
#     Aug-28-2026 preseason defense tier rankings -- see the module
#     docstring's DATA PROVENANCE note for the honest caveat (no free 2026
#     pass/run DVOA-style split exists publicly).
PASS_D_STRENGTH = {
    "BUF": 1, "MIN": 2, "CLE": 3, "NO": 4, "LAC": 5, "HOU": 6, "DEN": 7, "PHI": 8,
    "NE": 9, "SEA": 10, "GB": 11, "KC": 12, "ATL": 13, "LV": 14, "CAR": 15, "NYG": 16,
    "NYJ": 17, "MIA": 18, "LAR": 19, "DET": 20, "JAC": 21, "CHI": 22, "TEN": 23, "ARI": 24,
    "SF": 25, "CIN": 26, "TB": 27, "WAS": 28, "PIT": 29, "BAL": 30, "IND": 31, "DAL": 32,
}
RUSH_D_STRENGTH = {
    "JAC": 1, "DEN": 2, "SEA": 3, "HOU": 4, "TB": 5, "NE": 6, "IND": 7, "LAC": 8,
    "KC": 9, "BAL": 10, "SF": 11, "LAR": 12, "PIT": 13, "DET": 14, "TEN": 15, "CLE": 16,
    "LV": 17, "GB": 18, "NO": 19, "CAR": 20, "MIN": 21, "PHI": 22, "DAL": 23, "ATL": 24,
    "ARI": 25, "MIA": 26, "CHI": 27, "BUF": 28, "NYJ": 29, "WAS": 30, "NYG": 31, "CIN": 32,
}

def _build_early_sos_rank(defense_strength: dict) -> dict:
    """For every team, average the defensive strength of its 4 real Wk1-4
    opponents, then re-rank all 32 teams so 1 = easiest (faced the weakest
    defenses on average) .. 32 = hardest (faced the toughest)."""
    avg_opp_strength = {}
    for team, opponents in SCHEDULE_WK1_4.items():
        vals = [defense_strength[opp] for opp in opponents]
        avg_opp_strength[team] = sum(vals) / len(vals)
    # Higher average opponent-defense-rank number = weaker defenses faced =
    # easier slate, so sort descending and hand out rank 1 to the easiest.
    ordered = sorted(avg_opp_strength, key=lambda t: -avg_opp_strength[t])
    return {team: i + 1 for i, team in enumerate(ordered)}

SOS_PASSING_RANK = _build_early_sos_rank(PASS_D_STRENGTH)   # QB, WR, TE -- Wk1-4 only
SOS_RUSHING_RANK = _build_early_sos_rank(RUSH_D_STRENGTH)    # RB -- Wk1-4 only

def sos_tier(rank: int) -> int:
    if rank <= 6: return 1
    if rank <= 13: return 2
    if rank <= 19: return 3
    if rank <= 26: return 4
    return 5

SOS_LABELS = {1: "Very Soft", 2: "Soft", 3: "Neutral", 4: "Tough", 5: "Gauntlet"}

# ---- Real 2026 NFL bye weeks (verified against 2 independent sources: NFL.com's
#      own schedule-release article and Sports Illustrated, both cross-checked
#      for an exact match -- there is no Week 12 bye this season since every club
#      plays over the Thanksgiving slate; Week 11 has the most teams off at once). --
BYE_WEEKS = {
    "CAR": 5, "KC": 5,
    "CIN": 6, "DET": 6, "MIA": 6, "MIN": 6,
    "BUF": 7, "JAC": 7, "LAC": 7, "WAS": 7,
    "HOU": 8, "NO": 8, "NYG": 8, "SF": 8,
    "PIT": 9, "TEN": 9,
    "CHI": 10, "DEN": 10, "PHI": 10, "TB": 10,
    "ATL": 11, "CLE": 11, "GB": 11, "LAR": 11, "NE": 11, "SEA": 11,
    "BAL": 13, "IND": 13, "LV": 13, "NYJ": 13,
    "ARI": 14, "DAL": 14,
}
assert len(BYE_WEEKS) == 32, "BYE_WEEKS must cover all 32 NFL teams"

# =============================================================================
# 3. REAL FANTASYPROS ECR DATA (fetched 2026-08-29, Half-PPR Overall Draft
#    ECR, 108-expert consensus -- see the docstring for provenance/limits).
#    Format: (Name, Team, position-rank-number)  -- position-rank order IS
#    the ECR order, so list position doubles as the ECR sort.
# =============================================================================

ECR_QB = [
    ("Josh Allen","BUF"),("Lamar Jackson","BAL"),("Drake Maye","NE"),("Joe Burrow","CIN"),
    ("Jayden Daniels","WAS"),("Jalen Hurts","PHI"),("Caleb Williams","CHI"),("Justin Herbert","LAC"),
    ("Trevor Lawrence","JAC"),("Dak Prescott","DAL"),("Brock Purdy","SF"),("Jaxson Dart","NYG"),
    ("Bo Nix","DEN"),("Patrick Mahomes","KC"),("Matthew Stafford","LAR"),("Jared Goff","DET"),
    ("Jordan Love","GB"),("Baker Mayfield","TB"),("Kyler Murray","MIN"),("Tyler Shough","NO"),
    ("Malik Willis","MIA"),("Sam Darnold","SEA"),("C.J. Stroud","HOU"),("Cam Ward","TEN"),
    ("Daniel Jones","IND"),("Bryce Young","CAR"),("Jacoby Brissett","ARI"),("Geno Smith","NYJ"),
    ("Aaron Rodgers","PIT"),("Fernando Mendoza","LV"),("Tua Tagovailoa","ATL"),("Michael Penix Jr.","ATL"),
    ("Kirk Cousins","LV"),("Deshaun Watson","CLE"),("Shedeur Sanders","CLE"),("Carson Beck","ARI"),
    ("J.J. McCarthy","MIN"),("Mac Jones","SF"),("Justin Fields","KC"),("Ty Simpson","LAR"),
]

ECR_RB = [
    ("Jahmyr Gibbs","DET"),("Bijan Robinson","ATL"),("Christian McCaffrey","SF"),("Jonathan Taylor","IND"),
    ("James Cook","BUF"),("Chase Brown","CIN"),("Saquon Barkley","PHI"),("De'Von Achane","MIA"),
    ("Kenneth Walker III","KC"),("Omarion Hampton","LAC"),("Derrick Henry","BAL"),("Ashton Jeanty","LV"),
    ("Kyren Williams","LAR"),("Javonte Williams","DAL"),("Breece Hall","NYJ"),("Jeremiyah Love","ARI"),
    ("Josh Jacobs","GB"),("Travis Etienne","NO"),("D'Andre Swift","CHI"),("Cam Skattebo","NYG"),
    ("Bucky Irving","TB"),("Quinshon Judkins","CLE"),("David Montgomery","HOU"),("Bhayshul Tuten","JAC"),
    ("Jadarian Price","SEA"),("TreVeyon Henderson","NE"),("Rhamondre Stevenson","NE"),("Jaylen Warren","PIT"),
    ("Tony Pollard","TEN"),("Jonathon Brooks","CAR"),("Rico Dowdle","PIT"),("J.K. Dobbins","DEN"),
    ("Blake Corum","LAR"),("Chuba Hubbard","CAR"),("Jacory Croskey-Merritt","WAS"),("RJ Harvey","DEN"),
    ("Jordan Mason","MIN"),("Kenny Gainwell","TB"),("Kyle Monangai","CHI"),("Rachaad White","WAS"),
    ("Aaron Jones","MIN"),("Chris Rodriguez Jr.","JAC"),("Woody Marks","HOU"),("Tyler Allgeier","ARI"),
    ("Tyjae Spears","TEN"),("Keaton Mitchell","LAC"),("Tank Bigsby","PHI"),("Jonah Coleman","DEN"),
    ("Zach Charbonnet","SEA"),("Dylan Sampson","CLE"),("Isiah Pacheco","DET"),("MarShawn Lloyd","GB"),
    ("Alvin Kamara","NO"),("Tyrone Tracy Jr.","NYG"),("Brian Robinson","ATL"),("Braelon Allen","NYJ"),
    ("Mike Washington Jr.","LV"),("Ray Davis","BUF"),("Emmett Johnson","KC"),("Kimani Vidal","LAC"),
    ("Nicholas Singleton","TEN"),("Sean Tucker","TB"),("Jaydon Blue","DAL"),("Kaytron Allen","WAS"),
    ("George Holani","SEA"),("James Conner","ARI"),("Emanuel Wilson","SEA"),("Jaylen Wright","MIA"),
    ("Kaelon Black","SF"),("Justice Hill","BAL"),("Najee Harris","NYG"),("Chris Brooks","GB"),
    ("Jordan James","SF"),("Demond Claiborne","MIN"),("Samaje Perine","CIN"),("Isaiah Davis","NYJ"),
    ("Ollie Gordon II","MIA"),("Malik Davis","DAL"),("Ty Johnson","BUF"),("LeQuint Allen","JAC"),
    ("DJ Giddens","IND"),("Kendre Miller","NO"),("Seth McGowan","IND"),("Adam Randall","BAL"),
    ("Emari Demercado","KC"),("Devin Neal","NO"),("Brashard Smith","KC"),("Devin Singletary","NYG"),
    ("Trevor Etienne","CAR"),("Tahj Brooks","CIN"),("Trey Benson","ARI"),("Kaleb Johnson","PIT"),
    ("Jarquez Hunter","MIA"),("Jerome Ford","WAS"),("Isaac Guerendo","SF"),("Audric Estime","NO"),
    ("Jaleel McLaughlin","DEN"),("Will Shipley","PHI"),("Kareem Hunt","FA"),("Joe Mixon","FA"),
    ("Bam Knight","ARI"),("Eli Heidenreich","PIT"),("Jawhar Jordan","HOU"),("Michael Carter","TEN"),
    ("Raheim Sanders","CLE"),
]

ECR_WR = [
    ("Ja'Marr Chase","CIN"),("Puka Nacua","LAR"),("Jaxon Smith-Njigba","SEA"),("Amon-Ra St. Brown","DET"),
    ("CeeDee Lamb","DAL"),("Justin Jefferson","MIN"),("A.J. Brown","NE"),("Drake London","ATL"),
    ("Nico Collins","HOU"),("George Pickens","DAL"),("Chris Olave","NO"),("Malik Nabers","NYG"),
    ("DeVonta Smith","PHI"),("Rashee Rice","KC"),("Zay Flowers","BAL"),("Tee Higgins","CIN"),
    ("Ladd McConkey","LAC"),("Tetairoa McMillan","CAR"),("Jaylen Waddle","DEN"),("Garrett Wilson","NYJ"),
    ("Emeka Egbuka","TB"),("Terry McLaurin","WAS"),("Davante Adams","LAR"),("Luther Burden III","CHI"),
    ("Jameson Williams","DET"),("D.J. Moore","BUF"),("Christian Watson","GB"),("Rome Odunze","CHI"),
    ("Mike Evans","SF"),("Parker Washington","JAC"),("Marvin Harrison Jr.","ARI"),("Carnell Tate","TEN"),
    ("Brian Thomas Jr.","JAC"),("DK Metcalf","PIT"),("Chris Godwin","TB"),("Courtland Sutton","DEN"),
    ("Quentin Johnston","LAC"),("Michael Pittman Jr.","PIT"),("Michael Wilson","ARI"),("Alec Pierce","IND"),
    ("Josh Downs","IND"),("Wan'Dale Robinson","TEN"),("Stefon Diggs","WAS"),("Jordan Addison","MIN"),
    ("Jayden Reed","GB"),("Makai Lemon","PHI"),("Jakobi Meyers","JAC"),("Romeo Doubs","NE"),
    ("Xavier Worthy","KC"),("Jalen Coker","CAR"),("Matthew Golden","GB"),("Khalil Shakir","BUF"),
    ("KC Concepcion","CLE"),("Jordyn Tyson","NO"),("Deebo Samuel","SF"),("Rashid Shaheed","SEA"),
    ("De'Zhaun Stribling","SF"),("Adonai Mitchell","NYJ"),("Denzel Boston","CLE"),("Tre Tucker","LV"),
    ("Jerry Jeudy","CLE"),("Jalen McMillan","TB"),("Jauan Jennings","MIN"),("Tre' Harris","LAC"),
    ("Dontayvion Wicks","PHI"),("Kayshon Boutte","HOU"),("Ryan Flournoy","DAL"),("Pat Bryant","DEN"),
    ("Omar Cooper Jr.","NYJ"),("Malik Washington","MIA"),("Jalen Nailor","LV"),("Calvin Ridley","TEN"),
    ("Keenan Allen","IND"),("Jaylin Noel","HOU"),("Travis Hunter","JAC"),("Rashod Bateman","BAL"),
    ("Tank Dell","HOU"),("Cooper Kupp","SEA"),("Isaac TeSlaa","DET"),("Darnell Mooney","NYG"),
    ("Germie Bernard","PIT"),("Zachariah Branch","ATL"),("Troy Franklin","DEN"),("Devaughn Vele","NO"),
    ("Malachi Fields","NYG"),("Keon Coleman","BUF"),("Ja'Kobi Lane","BAL"),("Chris Bell","MIA"),
    ("Jack Bech","LV"),("Antonio Williams","WAS"),("Ted Hurst III","TB"),("Elic Ayomanor","TEN"),
    ("Tory Horton","SEA"),("Tyquan Thornton","KC"),("Caleb Douglas","MIA"),("Chimere Dike","TEN"),
    ("Darius Slayton","NYG"),("Christian Kirk","SF"),("Xavier Legette","CAR"),("Elijah Sarratt","BAL"),
    ("Marvin Mims","DEN"),("Cyrus Allen","KC"),("Kyle Williams","NE"),("Hollywood Brown","PHI"),
    ("Mack Hollins","NE"),("Isaiah Bond","CLE"),("Skyler Bell","BUF"),("Brandon Aiyuk","SF"),
    ("Demario Douglas","NE"),("Andrei Iosivas","CIN"),("Jahan Dotson","ATL"),("Tyreek Hill","FA"),
    ("Bryce Lance","NO"),("Jalen Tolbert","MIA"),("Xavier Hutchinson","HOU"),("Brenen Thompson","LAC"),
    ("Luke McCaffrey","WAS"),("Konata Mumpfield","LAR"),("Tez Johnson","TB"),("Kendrick Bourne","ARI"),
    ("Cedric Tillman","FA"),("Jalen Royals","KC"),("Joshua Palmer","BUF"),("Olamide Zaccheaus","ATL"),
    ("Treylon Burks","WAS"),("Chris Brazzell II","CAR"),("Malik Benson","LV"),("Savion Williams","GB"),
    ("KaVontae Turpin","DAL"),("Roman Wilson","PIT"),("Dont'e Thornton Jr.","LV"),
]

ECR_TE = [
    ("Brock Bowers","LV"),("Trey McBride","ARI"),("Colston Loveland","CHI"),("Tyler Warren","IND"),
    ("Tucker Kraft","GB"),("Harold Fannin","CLE"),("Sam LaPorta","DET"),("Kyle Pitts","ATL"),
    ("George Kittle","SF"),("Travis Kelce","KC"),("Dalton Kincaid","BUF"),("Dallas Goedert","PHI"),
    ("Isaiah Likely","NYG"),("Mark Andrews","BAL"),("Jake Ferguson","DAL"),("Juwan Johnson","NO"),
    ("Hunter Henry","NE"),("Chig Okonkwo","WAS"),("Dalton Schultz","HOU"),("Brenton Strange","JAC"),
    ("Terrance Ferguson","LAR"),("AJ Barner","SEA"),("T.J. Hockenson","MIN"),("Oronde Gadsden II","LAC"),
    ("Kenyon Sadiq","NYJ"),("Gunnar Helm","TEN"),("Pat Freiermuth","PIT"),("Cade Otton","TB"),
    ("David Njoku","LAC"),("Evan Engram","DEN"),("Colby Parkinson","LAR"),("Greg Dulcich","MIA"),
    ("Mason Taylor","NYJ"),("Theo Johnson","NYG"),("Eli Stowers","PHI"),("Mike Gesicki","CIN"),
    ("Darren Waller","CAR"),("Jake Tonges","SF"),("Darnell Washington","PIT"),("Oscar Delp","NO"),
    ("Michael Mayer","LV"),("Charlie Kolar","LAC"),("Elijah Arroyo","SEA"),("Erick All Jr.","CIN"),
    ("Eli Raridon","NE"),("Cole Kmet","CHI"),("Dawson Knox","BUF"),("Tyler Higbee","LAR"),
    ("Noah Gray","KC"),("Max Klare","LAR"),("Ja'Tavion Sanders","CAR"),("Justin Joly","DEN"),
]

# Superflex-specific Half-PPR Overall ECR -- only ranks 1-20 are publicly
# retrievable (FantasyPros paywalls the rest). Bonus context only.
SF_TOP20 = [
    ("Josh Allen", 1), ("Lamar Jackson", 2), ("Drake Maye", 3), ("Joe Burrow", 4),
    ("Jayden Daniels", 5), ("Jalen Hurts", 6), ("Jahmyr Gibbs", 7), ("Bijan Robinson", 8),
    ("Ja'Marr Chase", 9), ("Caleb Williams", 10), ("Puka Nacua", 11), ("Justin Herbert", 12),
    ("Jaxon Smith-Njigba", 13), ("Amon-Ra St. Brown", 14), ("Dak Prescott", 15),
    ("Trevor Lawrence", 16), ("Christian McCaffrey", 17), ("Jonathan Taylor", 18),
    ("James Cook", 19), ("CeeDee Lamb", 20),
]

# =============================================================================
# 3b. REAL ADP (Average Draft Position) -- FantasyPros Real-Time ADP,
#     Superflex view, Half-PPR, 12-team, fetched 2026-08-29 ("last updated
#     5 minutes ago" at fetch time). 250 real rows -- the tool caps there;
#     the ~78 deepest players in the 328-player pool have no published ADP
#     (expected -- they're below what any live draft room is actually
#     drafting). NOTE: this is 12-team ADP, not 10-team -- the RELATIVE
#     ORDER is still a good Superflex-aware proxy, but a literal "pick
#     47.6" won't map onto an exact 10-team snake slot. Confirmed genuinely
#     Superflex-formatted (QBs pulled way up vs. standard ADP -- Josh Allen
#     4.8 overall, Lamar Jackson 8.9) rather than standard/1-QB ADP.
#     Format: (Name, Team, ADP overall pick number).
# =============================================================================

ADP_DATA = [
    ("Bijan Robinson","ATL",3.3), ("Jahmyr Gibbs","DET",3.5), ("Josh Allen","BUF",4.8),
    ("Ja'Marr Chase","CIN",6.1), ("Puka Nacua","LAR",7.3), ("Jaxon Smith-Njigba","SEA",8.2),
    ("Christian McCaffrey","SF",8.6), ("Lamar Jackson","BAL",8.9), ("Amon-Ra St. Brown","DET",10.8),
    ("Jonathan Taylor","IND",11.2), ("James Cook III","BUF",13.3), ("Drake Maye","NE",14.4),
    ("CeeDee Lamb","DAL",15.3), ("Justin Jefferson","MIN",16.6), ("Saquon Barkley","PHI",17.3),
    ("Joe Burrow","CIN",18.8), ("De'Von Achane","MIA",20.0), ("Chase Brown","CIN",22.1),
    ("Jayden Daniels","WAS",23.2), ("Kenneth Walker III","KC",23.3), ("Jalen Hurts","PHI",25.1),
    ("Omarion Hampton","LAC",25.1), ("Ashton Jeanty","LV",25.7), ("Derrick Henry","BAL",25.8),
    ("A.J. Brown","NE",26.0), ("Drake London","ATL",26.4), ("Brock Bowers","LV",27.1),
    ("Caleb Williams","CHI",28.0), ("Nico Collins","HOU",28.8), ("George Pickens","DAL",31.1),
    ("Trey McBride","ARI",31.6), ("Malik Nabers","NYG",35.5), ("Justin Herbert","LAC",35.9),
    ("Chris Olave","NO",37.5), ("Jeremiyah Love","ARI",38.0), ("Kyren Williams","LAR",38.8),
    ("Dak Prescott","DAL",38.9), ("Rashee Rice","KC",40.8), ("DeVonta Smith","PHI",41.4),
    ("Trevor Lawrence","JAC",42.1), ("Breece Hall","NYJ",44.4), ("Colston Loveland","CHI",44.8),
    ("Zay Flowers","BAL",46.3), ("Javonte Williams","DAL",46.5), ("Tee Higgins","CIN",48.0),
    ("Tetairoa McMillan","CAR",48.9), ("Josh Jacobs","GB",49.7), ("Jaxson Dart","NYG",50.1),
    ("Ladd McConkey","LAC",50.1), ("Jaylen Waddle","DEN",50.5), ("Bo Nix","DEN",51.4),
    ("Travis Etienne Jr.","NO",54.2), ("Garrett Wilson","NYJ",54.2), ("Patrick Mahomes II","KC",54.7),
    ("Emeka Egbuka","TB",55.6), ("Brock Purdy","SF",57.1), ("Cam Skattebo","NYG",57.7),
    ("D'Andre Swift","CHI",58.5), ("Matthew Stafford","LAR",58.7), ("Tyler Warren","IND",58.7),
    ("Bucky Irving","TB",61.7), ("Luther Burden III","CHI",64.1), ("DJ Moore","BUF",64.4),
    ("Jared Goff","DET",64.5), ("Terry McLaurin","WAS",64.5), ("Quinshon Judkins","CLE",67.4),
    ("Davante Adams","LAR",67.8), ("David Montgomery","HOU",68.2), ("Jameson Williams","DET",72.3),
    ("Bhayshul Tuten","JAC",72.9), ("Rome Odunze","CHI",75.1), ("Jadarian Price","SEA",75.5),
    ("Sam LaPorta","DET",75.9), ("Jordan Love","GB",77.0), ("Kyler Murray","MIN",77.3),
    ("TreVeyon Henderson","NE",77.4), ("Tucker Kraft","GB",77.5), ("Christian Watson","GB",78.2),
    ("Mike Evans","SF",78.4), ("Parker Washington","JAC",78.9), ("Baker Mayfield","TB",80.9),
    ("Harold Fannin Jr.","CLE",81.4), ("Carnell Tate","TEN",82.7), ("Jaylen Warren","PIT",83.2),
    ("Kyle Pitts Sr.","ATL",85.1), ("Tyler Shough","NO",87.0), ("Marvin Harrison Jr.","ARI",87.4),
    ("Rhamondre Stevenson","NE",87.6), ("Brian Thomas Jr.","JAC",92.0), ("DK Metcalf","PIT",93.0),
    ("Sam Darnold","SEA",96.1), ("Tony Pollard","TEN",97.2), ("Rico Dowdle","PIT",99.7),
    ("George Kittle","SF",100.7), ("Travis Kelce","KC",100.9), ("Chris Godwin Jr.","TB",101.3),
    ("Jonathon Brooks","CAR",101.4), ("Courtland Sutton","DEN",101.5), ("Michael Wilson","ARI",102.3),
    ("RJ Harvey","DEN",103.3), ("J.K. Dobbins","DEN",104.9), ("Dalton Kincaid","BUF",105.9),
    ("Malik Willis","MIA",106.6), ("C.J. Stroud","HOU",111.0), ("Blake Corum","LAR",111.4),
    ("Chuba Hubbard","CAR",112.0), ("Jayden Reed","GB",112.1), ("Stefon Diggs","WAS",112.4),
    ("Michael Pittman Jr.","PIT",112.5), ("Alec Pierce","IND",112.6), ("Jordan Mason","MIN",113.0),
    ("Makai Lemon","PHI",113.3), ("Isaiah Likely","NYG",114.3), ("Daniel Jones","IND",114.7),
    ("Jordan Addison","MIN",116.3), ("Josh Downs","IND",117.5), ("Jacory Croskey-Merritt","WAS",117.9),
    ("Kenny Gainwell","TB",119.6), ("De'Zhaun Stribling","SF",120.9), ("Quentin Johnston","LAC",121.3),
    ("Kyle Monangai","CHI",121.7), ("Cam Ward","TEN",121.7), ("Wan'Dale Robinson","TEN",124.4),
    ("Jake Ferguson","DAL",125.2), ("Matthew Golden","GB",125.4), ("KC Concepcion","CLE",125.4),
    ("Dallas Goedert","PHI",125.8), ("Mark Andrews","BAL",126.2), ("Bryce Young","CAR",126.9),
    ("Jordyn Tyson","NO",130.1), ("Aaron Jones Sr.","MIN",130.7), ("Fernando Mendoza","LV",134.6),
    ("Rachaad White","WAS",135.7), ("Jakobi Meyers","JAC",136.2), ("Xavier Worthy","KC",140.0),
    ("Romeo Doubs","NE",140.3), ("Deebo Samuel Sr.","SF",141.2), ("Jalen Coker","CAR",142.0),
    ("Chris Rodriguez Jr.","JAC",142.9), ("Aaron Rodgers","PIT",143.3), ("Khalil Shakir","BUF",144.2),
    ("Woody Marks","HOU",144.2), ("Tyler Allgeier","ARI",145.0), ("Mike Washington Jr.","LV",145.1),
    ("MarShawn Lloyd","GB",145.6), ("Jacoby Brissett","ARI",145.9), ("Geno Smith","NYJ",147.6),
    ("Rashid Shaheed","SEA",148.6), ("Hunter Henry","NE",148.6), ("Juwan Johnson","NO",148.9),
    ("Zach Charbonnet","SEA",151.2), ("Brenton Strange","JAC",153.0), ("Tyjae Spears","TEN",153.4),
    ("Jonah Coleman","DEN",154.6), ("Brian Robinson Jr.","ATL",156.6), ("Keaton Mitchell","LAC",158.4),
    ("Ja'Kobi Lane","BAL",158.5), ("Tyrone Tracy Jr.","NYG",158.6), ("Denzel Boston","CLE",158.9),
    ("Tank Bigsby","PHI",159.1), ("Dalton Schultz","HOU",160.2), ("Chig Okonkwo","WAS",161.8),
    ("Jalen Nailor","LV",162.2), ("Oronde Gadsden II","LAC",162.2), ("Alvin Kamara","NO",162.5),
    ("Keenan Allen","IND",163.6), ("Dylan Sampson","CLE",163.7), ("Tre Tucker","LV",163.7),
    ("Tua Tagovailoa","ATL",164.1), ("T.J. Hockenson","MIN",164.4), ("Jalen McMillan","TB",164.5),
    ("Kenyon Sadiq","NYJ",164.8), ("Deshaun Watson","CLE",165.5), ("Kirk Cousins","LV",165.9),
    ("Cyrus Allen","KC",165.9), ("Isiah Pacheco","DET",166.5), ("Malik Washington","MIA",166.5),
    ("Kaelon Black","SF",166.8), ("Braelon Allen","NYJ",167.5), ("Kayshon Boutte","HOU",167.7),
    ("Travis Hunter","JAC",168.0), ("Michael Penix Jr.","ATL",168.1), ("Emmett Johnson","KC",168.1),
    ("AJ Barner","SEA",168.1), ("Jerry Jeudy","CLE",168.3), ("Shedeur Sanders","CLE",169.1),
    ("Omar Cooper Jr.","NYJ",169.3), ("Tank Dell","HOU",169.6), ("Jauan Jennings","MIN",170.4),
    ("Devaughn Vele","NO",170.7), ("Caleb Douglas","MIA",171.1), ("Malachi Fields","NYG",171.3),
    ("Dontayvion Wicks","PHI",171.3), ("Adonai Mitchell","NYJ",171.3), ("James Conner","ARI",171.5),
    ("Terrance Ferguson","LAR",171.9), ("Pat Bryant","DEN",172.1), ("Justice Hill","BAL",172.1),
    ("Zachariah Branch","ATL",172.5), ("Cooper Kupp","SEA",173.3), ("Isaac TeSlaa","DET",173.6),
    ("Jaydon Blue","DAL",173.7), ("Kaytron Allen","WAS",173.7), ("Ray Davis","BUF",174.2),
    ("Colby Parkinson","LAR",174.4), ("Eli Stowers","PHI",174.4), ("Troy Franklin","DEN",174.4),
    ("Antonio Williams","WAS",174.4), ("Tyreek Hill","FA",174.5), ("Najee Harris","NYG",174.5),
    ("Calvin Ridley","TEN",174.7), ("Nicholas Singleton","TEN",174.9), ("Tory Horton","SEA",175.1),
    ("Tre' Harris","LAC",175.3), ("Rashod Bateman","BAL",175.3), ("George Holani","SEA",175.4),
    ("Chris Bell","MIA",175.5), ("Jayden Higgins","HOU",175.6), ("Demond Claiborne","MIN",175.8),
    ("Keon Coleman","BUF",175.8), ("Gunnar Helm","TEN",175.9), ("Ryan Flournoy","DAL",175.9),
    ("Jordan James","SF",176.3), ("Jake Tonges","SF",176.5), ("Kimani Vidal","LAC",176.6),
    ("Cade Otton","TB",176.6), ("David Njoku","LAC",176.6), ("Greg Dulcich","MIA",176.6),
    ("Evan Engram","DEN",176.8), ("Kendre Miller","NO",176.9), ("Darren Waller","CAR",176.9),
    ("Carson Beck","ARI",177.0), ("Ty Simpson","LAR",177.1), ("Sean Tucker","TB",177.3),
    ("Justin Fields","KC",177.3), ("Pat Freiermuth","PIT",177.4), ("J.J. McCarthy","MIN",178.0),
    ("Jaylin Noel","HOU",178.6), ("Samaje Perine","CIN",178.9), ("Trey Benson","ARI",179.6),
    ("Chimere Dike","TEN",179.7), ("Jaylen Wright","MIA",180.0), ("Mac Jones","SF",180.3),
    ("Darnell Mooney","NYG",181.7),
]

# =============================================================================
# 4. ENRICHMENT -- prior situational research merged in by (normalized) name.
#    (ContractYear Y/N, Flag, Notes). Only players actually researched the
#    week of Aug 24-29, 2026 get non-empty Flag/Notes -- everyone else in
#    the 328-player ECR pool gets a generic, honest placeholder note.
# =============================================================================

ENRICHMENT = {
    "josh allen": ("N", "STUD", "Unanimous 2026 QB1; top-2 fantasy finish 5 straight years"),
    "lamar jackson": ("N", "STUD", "Elite dual-threat despite a down 2025; SF Overall #2"),
    "drake maye": ("N", "", "Superflex 1st-round arm; SF Overall #3"),
    "joe burrow": ("N", "", "Superflex 1st-round arm talent; SF Overall #4"),
    "jared goff": ("N", "FLOOR", "4 straight 4,400+ yd seasons; historically underrated in ADP"),
    "jalen hurts": ("N", "BUST", "Rushing attempts down 3 straight yrs; tush-push success rate 90%->60%"),
    "jayden daniels": ("N", "GAMBLE", "Missed 10 games in 2025; new under-center scheme; boom/bust"),
    "caleb williams": ("N", "", "Superflex upside arm; SF Overall #10"),
    "justin herbert": ("N", "", "New OC Mike McDaniel arrives in LAC"),
    "trevor lawrence": ("N", "", "UPDATED 8/30: fully healthy entering 2026 (Dec 2025 shoulder/AC-joint surgery is resolved), "
                        "confirmed Week 1 starter per camp reports -- Superflex-relevant arm, needs weapons to step up"),
    "brock purdy": ("N", "", "Efficient offense, Mike Evans now a weapon"),
    "dak prescott": ("N", "", "Steady superflex QB1/2 floor"),
    "jaxson dart": ("N", "BUST", "15 total TD in 12 rookie starts; new OC installing a different scheme"),
    "patrick mahomes": ("N", "GAMBLE", "ACL-scare discount; buy-low if medicals check out"),
    "kyler murray": ("N", "SLEEPER", "Won the Vikings' Wk1 job over McCarthy; elite weapons, still-depressed price"),
    "tyler shough": ("N", "SLEEPER", "Kellen Moore's offense; deep-league Superflex QB2 stash"),
    "malik willis": ("N", "SLEEPER", "Won Miami's job under an entirely new coaching staff; rushing ceiling"),
    "baker mayfield": ("Y", "DEPTH", "Contract year; real starting job in Tampa Bay"),
    "tua tagovailoa": ("Y", "DEPTH", "Contract year; now Atlanta's QB1 by depth chart, unproven fit"),
    "anthony richardson": ("Y", "DEPTH", "Contract year; backup behind Daniel Jones in Indianapolis"),
    "justin fields": ("Y", "DEPTH", "Contract year; rushing-only streamer if he starts in KC"),
    "jacoby brissett": ("Y", "DEPTH", "Contract year; Arizona's QB1 but a pure streamer"),
    "bryce young": ("N", "", "NEW 8/30: confirmed Panthers starter, strong/positive camp reports (coach cites 'ownership and "
                    "mastery of the concepts'), no benching threat found -- the 2024 benching is old news, not current"),
    "deshaun watson": ("Y", "DEPTH", "Contract year; Cleveland's QB1, name value only"),
    "aaron rodgers": ("Y", "DEPTH", "Contract year; age 42, late-round streamer only"),
    "kirk cousins": ("Y", "DEPTH", "Contract year; age 38, backup in Las Vegas"),
    "davis mills": ("Y", "DEPTH", "Contract year; backup to Stroud"),
    "mac jones": ("Y", "DEPTH", "Contract year; backup to Purdy"),
    "marcus mariota": ("Y", "DEPTH", "Contract year; backup to Daniels (WAS-adjacent depth)"),
    "will levis": ("Y", "DEPTH", "Contract year; QB depth behind Ward"),
    "aidan o'connell": ("Y", "DEPTH", "Contract year; depth arm"),

    "jahmyr gibbs": ("N", "STUD", "Top-of-market receiving-back profile; SF Overall #7. UPDATED 8/30: hamstring issue after "
                     "signing his Aug extension pulled him from practice -- Campbell says 'on schedule' for the opener, "
                     "confirm his practice status the morning of your draft"),
    "bijan robinson": ("N", "STUD", "Bell-cow touches plus real receiving work; SF Overall #8"),
    "christian mccaffrey": ("N", "BUST", "Age 30; heavy-touch seasons have twice preceded serious injury years"),
    "jonathan taylor": ("N", "STUD", "Led the NFL in snap share (89%) in 2025; SF Overall #18"),
    "chase brown": ("Y", "FLOOR", "Contract year; cited as the predictable alternative to a risky McCaffrey pick"),
    "james cook": ("N", "FLOOR", "Steady early-down/receiving mix; SF Overall #19"),
    "de'von achane": ("N", "BUST", "McDaniel departure + unproven Willis at QB let defenses key on him"),
    "saquon barkley": ("N", "SLEEPER", "2025 dip looks like a blocking/scheme issue; new OC brings a livelier scheme"),
    "javonte williams": ("N", "BUST", "Post-bye decline 5.2->4.3 YPC; bottom 12% in receiving yards per route run"),
    "bucky irving": ("N", "FLOOR", "UPDATED 8/30: fully cleared from offseason shoulder surgery, 'full-go' and 'unquestioned starter "
                     "from Day One' per camp reports -- White is gone (now Washington), no more committee threat"),
    "bhayshul tuten": ("N", "BUST", "Averaged just 3.7 YPC; LeQuint Allen handles the receiving work"),
    "jadarian price": ("N", "GAMBLE", "1st-round rookie capital into an unresolved timeshare with Charbonnet"),
    "trevveyon henderson": ("N", "BUST", "Bottom 15% yards after contact; Rhamondre Stevenson took over the lead role late"),
    "tony pollard": ("Y", "DEPTH", "Contract year; Titans' clear RB1"),
    "jonathon brooks": ("N", "BUST", "9 career carries and a second ACL tear; Chuba Hubbard remains the starter"),
    "rico dowdle": ("N", "BUST", "Collapsed over final 9 games (3.2 YPC) as the lead-ish back -- also Warren's direct handcuff"),
    "blake corum": ("N", "HANDCUFF", "Backs up Kyren Williams after a 5.1 YPC season"),
    "kyle monangai": ("N", "HANDCUFF", "Backs up D'Andre Swift; strong rookie receiving role, efficiency dipped late"),
    "rachaad white": ("Y", "HANDCUFF", "Contract year; insurance behind Jacory Croskey-Merritt"),
    "aaron jones": ("Y", "DEPTH", "Contract year; Vikings committee back"),
    "tyler allgeier": ("N", "GAMBLE", "UPDATED 8/30: signed with Arizona into a crowded 3-way mix with rookie Jeremiyah Love and "
                       "James Conner -- one early depth chart even lists him RB1, but treat the touch split as unsettled"),
    "tyjae spears": ("Y", "DEPTH", "Contract year; Pollard's handcuff in Tennessee"),
    "keaton mitchell": ("N", "HANDCUFF", "Backs up Omarion Hampton; new OC McDaniel scheme fit"),
    "tank bigsby": ("Y", "DEPTH", "Contract year; Eagles RB2 behind Barkley"),
    "jonah coleman": ("N", "HANDCUFF", "UPDATED 8/30: RB3 behind J.K. Dobbins AND RJ Harvey, but specifically pegged to inherit the "
                      "3rd-down passing/pass-block role -- a real if limited-touch role, not a pure emergency stash"),
    "zach charbonnet": ("Y", "DEPTH", "Contract year; recovering from an ACL tear, timeshare with rookie Price unresolved"),
    "isiah pacheco": ("Y", "DEPTH", "Contract year; Lions RB2"),
    "marshawn lloyd": ("N", "HANDCUFF", "Josh Jacobs officially charged, suspension likely -- direct beneficiary"),
    "alvin kamara": ("Y", "DEPTH", "Contract year; still New Orleans' starter"),
    "brian robinson": ("Y", "DEPTH", "Contract year; Bijan's direct handcuff in Atlanta"),
    "james conner": ("Y", "DEPTH", "Contract year; Arizona's early-down/goal-line starter"),
    "emanuel wilson": ("Y", "DEPTH", "Contract year; Seattle depth behind the Price/Charbonnet battle"),
    "kaelon black": ("N", "HANDCUFF", "Competing for McCaffrey's direct backup job"),
    "justice hill": ("Y", "DEPTH", "Contract year; Ravens passing-down back"),
    "najee harris": ("Y", "DEPTH", "Contract year; Giants early-down back"),
    "jordan james": ("N", "HANDCUFF", "Competing for McCaffrey's direct backup job"),
    "samaje perine": ("Y", "DEPTH", "Contract year; Bengals passing-down depth"),
    "ty johnson": ("Y", "DEPTH", "Contract year; Bills RB depth"),
    "kendre miller": ("Y", "DEPTH", "Contract year; Kamara's handcuff in New Orleans"),
    "isaac guerendo": ("N", "HANDCUFF", "Competing for McCaffrey's direct backup job"),
    "rhamondre stevenson": ("N", "HANDCUFF", "NEW 8/30: genuine committee with 2nd-year TreVeyon Henderson, not a clear lead back -- "
                            "touches likely split all season, treat as backend RB depth not a locked-in starter"),
    "david montgomery": ("N", "GAMBLE", "NEW 8/30: traded from Detroit to Houston (Gibbs is now the Lions' unquestioned bell-cow); "
                         "lands in a crowded 3-way Texans mix with Woody Marks and Nick Chubb, no role locked down"),
    "kenny gainwell": ("N", "HANDCUFF", "NEW 8/30: Bowles calls him Tampa's 'RB1-B' -- real passing-down role and the direct "
                       "beneficiary if Irving's recently-surgically-repaired shoulder has any setback"),
    "breece hall": ("N", "GAMBLE", "NEW 8/30: groin strain in Aug camp (out 8/17); team expects a Week 1 return but this is worth "
                    "confirming the morning of your draft, not assuming"),

    "ceedee lamb": ("N", "STUD", "Safest true WR1 anchor; elite target volume"),
    "ja'marr chase": ("N", "STUD", "Consensus best WR in fantasy; SF Overall #9"),
    "amon-ra st. brown": ("N", "STUD", "\"A sure thing\" in Round 1 again; SF Overall #14"),
    "justin jefferson": ("N", "STUD", "2025 dip was QB-driven (McCarthy); Murray at QB is a real bounce-back thesis"),
    "puka nacua": ("Y", "GAMBLE", "Contract year; active psoas injury since mid-Aug; SF Overall #11 despite the risk"),
    "a.j. brown": ("Y", "", "Contract year; new WR1 in New England after leaving Philadelphia"),
    "d.j. moore": ("N", "FLOOR", "Confirmed move to Buffalo; reported as Josh Allen's \"de facto first read\""),
    "zay flowers": ("N", "FLOOR", "Efficient, high-floor value target"),
    "devonta smith": ("N", "FLOOR", "Steady target share in an efficient offense"),
    "tee higgins": ("N", "BUST", "Never topped 1,100 yards; missed 12 games over 3 yrs incl. 2 concussions in 2025"),
    "rashee rice": ("N", "GAMBLE", "Top-15 talent; real suspension/probation uncertainty"),
    "malik nabers": ("N", "GAMBLE", "Recovering well from a 2025 ACL tear, trending to a Wk1 return; price hasn't fully caught up"),
    "keenan allen": ("Y", "FLOOR", "Contract year; still commands 120+ targets in a crowded room"),
    "davante adams": ("Y", "BUST", "Contract year; turns 34 in Dec, snaps already shrinking pre-injury"),
    "jameson williams": ("N", "BUST", "Both breakout years required another player's absence; 3rd option when healthy"),
    "khalil shakir": ("N", "BUST", "Snap rate fell 76%->57%; Moore's arrival compresses his target share further"),
    "mike evans": ("Y", "BUST", "Contract year; new team/unproven QB fit; played 17/14/8 games last 3 yrs"),
    "xavier worthy": ("N", "SLEEPER", "Only 23; KC lacks depth behind Rice, value rises if Rice sits"),
    "george pickens": ("Y", "DEPTH", "Contract year; real starter-caliber weapon in Dallas"),
    "chris godwin": ("Y", "DEPTH", "Contract year; steady Tampa Bay starter"),
    "jordan addison": ("Y", "DEPTH", "Contract year; real WR2 in a good Vikings offense"),
    "jauan jennings": ("Y", "BUST", "Contract year; excellent blocker but caps his own target share behind Addison"),
    "josh downs": ("Y", "DEPTH", "Contract year; Colts slot option"),
    "darnell mooney": ("Y", "DEPTH", "Contract year; Giants WR2"),
    "deebo samuel": ("Y", "DEPTH", "Contract year; multi-purpose 49ers weapon"),
    "stefon diggs": ("N", "DEPTH", "Washington WR2, veteran"),
    "jalen coker": ("Y", "SLEEPER", "UPDATED 8/30: locked in as a starting OUTSIDE WR next to McMillan (not a slot role as "
                    "previously scouted) -- only the WR3 job behind them is contested"),
    "tre tucker": ("Y", "SLEEPER", "Contract year; locked in as Raiders' WR1 at a conservative price"),
    "courtland sutton": ("N", "BUST", "Jaylen Waddle's arrival mirrors past attempts to feature other WRs; turns 31"),
    "matthew golden": ("N", "SLEEPER", "UPDATED 8/30: thin rookie-year production (29/361/0), but camp reports now have him as "
                       "GB's clear #2/#3 target now that Doubs is gone (signed with NE) -- real opportunity behind Watson"),
    "chris bell": ("N", "SLEEPER", "Post-ACL; could be Miami's top size/speed WR if healthy"),
    "kayshon boutte": ("Y", "DEPTH", "Contract year; Texans depth"),
    "tank dell": ("Y", "DEPTH", "Contract year; Texans depth, buried in a crowded room"),
    "quentin johnston": ("Y", "DEPTH", "Contract year; Chargers WR3"),
    "christian kirk": ("Y", "DEPTH", "Contract year; 49ers depth"),
    "marvin mims": ("Y", "DEPTH", "Contract year; Broncos big-play role"),
    "cedric tillman": ("Y", "DEPTH", "CAUTION: previously scouted as Browns WR6 (contract year); FantasyPros' fresh Aug-29 pull "
                        "tags him FA (unsigned) -- verify his actual roster status before drafting."),
    "jalen tolbert": ("Y", "DEPTH", "Contract year; Dolphins depth-chart WR1"),
    "parker washington": ("Y", "DEPTH", "Contract year; Jaguars depth-chart WR1"),
    "tre' harris": ("N", "BUST", "Role fully dependent on Quentin Johnston's absence"),
    "jordyn tyson": ("N", "BUST", "Injured in all 4 college seasons plus a May injury"),
    "alec pierce": ("N", "BUST", "Ankle injury complications, no clear return timeline"),
    "ja'kobi lane": ("N", "BUST", "Camp buzz doesn't guarantee a role behind Flowers/Andrews"),
    "michael wilson": ("N", "BUST", "New OC shifts scheme; 3rd option behind Harrison Jr. and McBride"),
    "treylon burks": ("Y", "DEPTH", "Contract year; Commanders depth"),
    "demario douglas": ("Y", "DEPTH", "Contract year; Patriots depth"),
    "ryan flournoy": ("Y", "DEPTH", "Contract year; Cowboys depth"),
    "xavier hutchinson": ("Y", "DEPTH", "Contract year; Texans depth"),
    "andrei iosivas": ("Y", "DEPTH", "Contract year; Bengals depth"),
    "tyreek hill": ("N", "BUST", "Severe 2025 knee injury (dislocation + multi-ligament); unsigned FA; real chance he's done -- avoid entirely"),
    "chris olave": ("N", "GAMBLE", "NEW 8/30: well-documented multi-concussion history (his new contract even has a guarantee "
                    "waiver for a career-ending concussion) -- had an Aug practice scare, reportedly fine, but recurrence risk is real"),
    "brian thomas": ("N", "SLEEPER", "NEW 8/30: 'career-best shape' per camp reports, primed for Jacksonville's true WR1 role "
                     "-- projected ~22-25% target share, best big-play traits in a crowded room"),
    "jakobi meyers": ("N", "FLOOR", "NEW 8/30: traded from Las Vegas to Jacksonville; profiles as JAX's slot/possession weapon "
                      "(~22-25% target share) alongside Thomas and a limited-role Travis Hunter"),
    "michael pittman": ("Y", "FLOOR", "NEW 8/30: traded from Indianapolis to Pittsburgh, signed a 3yr/$59M extension there -- "
                        "presumptive #2 behind DK Metcalf, but capped ceiling in a low-volume passing offense"),
    "rashid shaheed": ("N", "SLEEPER", "NEW 8/30: healthy, no suspension found, viewed as a real breakout/stash candidate who "
                       "may also retain punt/kick return duties for added floor"),
    "jaylen waddle": ("N", "", "NEW 8/30: traded from Miami to Denver this offseason -- treated as a value WR1-upside play in "
                      "Bo Nix's offense, not a Dolphin any longer"),
    "romeo doubs": ("N", "", "NEW 8/30: signed with New England after Green Bay passed on re-signing him (rookie Matthew Golden "
                    "was ready to inherit the role) -- early-down/possession role with the Patriots now"),

    "brock bowers": ("N", "STUD", "1.0 TE premium turns him into a WR1-scoring asset"),
    "trey mcbride": ("N", "BUST", "2025 volume inflated by Arizona playing from behind all year; priced for a repeat"),
    "sam laporta": ("Y", "DEPTH", "Contract year; real difference-making receiving TE1"),
    "tucker kraft": ("Y", "DEPTH", "Contract year; ascending real TE1 target"),
    "harold fannin": ("N", "BUST", "Benefited from Njoku's injury; new additions should restore balanced play-calling"),
    "dallas goedert": ("Y", "FLOOR", "Contract year; TE4 with 11 TDs in 2025, more volume now that A.J. Brown is gone. "
                       "UPDATED 8/30: rookie TE Eli Stowers (2nd-rounder) + added WR depth threaten his target share -- one "
                       "analyst projects him closer to TE15 than TE10, still startable but trending down not up"),
    "george kittle": ("N", "GAMBLE", "Torn Achilles in the 2025 finale but a fast-recovery location; priced below his TE3 talent level"),
    "isaiah likely": ("N", "SLEEPER", "3yr/$40M deal, reunited with new Giants HC Harbaugh; Dart's 2nd read"),
    "t.j. hockenson": ("Y", "DEPTH", "Contract year; real receiving TE1 when healthy"),
    "kenyon sadiq": ("N", "BUST", "Sports-hernia recovery ongoing while Mason Taylor gains reps"),
    "dalton schultz": ("N", "SLEEPER", "Projected 100+ targets again as Stroud's clear No. 2"),
    "evan engram": ("Y", "DEPTH", "Contract year; Broncos starting TE"),
    "colby parkinson": ("Y", "DEPTH", "Contract year; Rams TE2"),
    "greg dulcich": ("Y", "DEPTH", "Contract year; Dolphins starting TE"),
    "michael mayer": ("Y", "DEPTH", "Contract year; Raiders TE2"),
    "darnell washington": ("Y", "DEPTH", "Contract year; Steelers blocking-TE2"),
    "brenton strange": ("Y", "DEPTH", "Contract year; Jaguars starting TE"),
    "kyle pitts": ("N", "BUST", "Averaged just 9.6 PPG in games Drake London also played"),
}

# =============================================================================
# 5. NAME NORMALIZATION (handles Jr./Sr./II/III/IV + punctuation drift so the
#    ENRICHMENT dict above matches FantasyPros' occasionally-fuller names)
# =============================================================================

_SUFFIX_RE = re.compile(r"\s+(jr\.?|sr\.?|ii|iii|iv|v)$", re.IGNORECASE)

def normalize_name(name: str) -> str:
    n = name.lower().replace(".", "").replace("'", "").replace("’", "")
    n = _SUFFIX_RE.sub("", n).strip()
    n = _SUFFIX_RE.sub("", n).strip()  # handles "Patrick Mahomes II" -> already fine, but double suffix safety
    return n

# =============================================================================
# 6. BUILD THE MASTER DATAFRAME
# =============================================================================

def build_pos_df(pos, ecr_list):
    rows = []
    for i, (name, team) in enumerate(ecr_list, start=1):
        key = normalize_name(name)
        cy, flag, note = ENRICHMENT.get(key, ("N", "", "ECR-ranked; no dedicated situational research yet -- "
                                                          "verify news/depth chart closer to draft time."))
        rows.append({"Name": name, "Team": team, "Pos": pos, "ECR_Pos": i, "ContractYear": cy,
                     "Flag": flag, "Notes": note})
    return pd.DataFrame(rows)

df = pd.concat([
    build_pos_df("QB", ECR_QB),
    build_pos_df("RB", ECR_RB),
    build_pos_df("WR", ECR_WR),
    build_pos_df("TE", ECR_TE),
], ignore_index=True)

# ---- generic data-driven tiering: biggest gaps on ANY sorted numeric column
#      become tier walls. Used twice below: once per-position on ECR_Pos (for
#      the reference "POS TIER" column), and once globally on GlobalValue
#      (for the primary "TIER" column that drives the whole board's layout).

def compute_tiers_generic(sub_df: pd.DataFrame, value_col: str, ascending: bool, n_tiers: int):
    """Returns (sub_df with a '_tier_tmp' column, cliff_tier_number). cliff_tier_number
    is the tier whose LAST row sits right before the single biggest gap in the
    sorted value_col -- i.e. the tier to flag "grab someone here, values crash
    right after this" (matches this kit's existing cliff-banner convention:
    the warning shows on the tier BEFORE the drop, not after it)."""
    sub_df = sub_df.sort_values(value_col, ascending=ascending).reset_index(drop=True)
    vals = sub_df[value_col].tolist()
    gaps = [(abs(vals[i+1] - vals[i]), i) for i in range(len(vals) - 1)]
    candidates = sorted([g for g in gaps if 0 < g[1] < len(vals) - 2], key=lambda x: -x[0])
    chosen = []
    min_spacing = max(2, len(vals) // 20)
    for gap_size, idx in candidates:
        if all(abs(idx - c[1]) >= min_spacing for c in chosen):
            chosen.append((gap_size, idx))
        if len(chosen) == n_tiers - 1:
            break
    chosen.sort(key=lambda x: x[1])
    max_gap_idx = max(chosen, key=lambda x: x[0])[1] if chosen else None

    tiers = [1] * len(vals)
    boundary_set = {idx for _, idx in chosen}
    t = 1
    for i in range(len(vals)):
        tiers[i] = t
        if i in boundary_set:
            t += 1
    sub_df["_tier_tmp"] = tiers
    cliff_tier_number = tiers[max_gap_idx] if max_gap_idx is not None else None
    return sub_df, cliff_tier_number

# ---- 6a. Per-position POS TIER (reference column, also drives the SF/TEP tag) --
pos_tiered_frames = []
for pos in ["QB", "RB", "WR", "TE"]:
    sub, _ = compute_tiers_generic(df[df["Pos"] == pos], "ECR_Pos", ascending=True, n_tiers=N_TIERS)
    sub = sub.rename(columns={"_tier_tmp": "PosTier"})
    pos_tiered_frames.append(sub)
df = pd.concat(pos_tiered_frames, ignore_index=True)

# ---- SoS (real Weeks 1-4 isolated metric -- see section 2 above) ----
def _sos_rank(row):
    table = SOS_RUSHING_RANK if row["Pos"] == "RB" else SOS_PASSING_RANK
    return table.get(row["Team"], 16)

df["Bye"] = df["Team"].apply(lambda t: BYE_WEEKS.get(t, None))  # None/blank for the FA row (Cedric Tillman)
df["SoS_Rank"] = df.apply(_sos_rank, axis=1)
df["SoS_Tier"] = df.apply(lambda r: 3 if r["Team"] == "FA" else sos_tier(r["SoS_Rank"]), axis=1)
df["SoS_Label"] = df.apply(lambda r: "N/A (FA)" if r["Team"] == "FA" else SOS_LABELS[r["SoS_Tier"]], axis=1)

# ---- Superflex bonus tag + SF/TEP modifier (uses PosTier, not the global Tier,
#      since "is this an elite QB/TE" is inherently a within-position question) --
SF_TOP20_MAP = {normalize_name(n): rank for n, rank in SF_TOP20}

def _sf_tep_tag(row):
    key = normalize_name(row["Name"])
    sf_rank = SF_TOP20_MAP.get(key)
    base = ""
    if row["Pos"] == "QB":
        base = "Superflex Elite" if row["PosTier"] <= 2 else ("Superflex Starter" if row["PosTier"] <= 4 else "Superflex Depth")
    elif row["Pos"] == "TE":
        base = "TEP Cheat Code" if row["PosTier"] <= 2 else ("TEP Boost" if row["PosTier"] <= 4 else "")
    if sf_rank:
        base = (base + f" | SF Overall #{sf_rank}").strip(" |")
    return base

df["SF_TEP_Tag"] = df.apply(_sf_tep_tag, axis=1)

# =============================================================================
# 7. GLOBAL VALUE MODEL (Value-Based Drafting)
#    Converts real ECR rank -> an estimated points curve per position ->
#    subtracts a replacement-level baseline -> a single cross-positional
#    "GlobalValue" scale that QB/RB/WR/TE can be sorted and tiered on
#    TOGETHER. This is what actually answers "who's the better player
#    regardless of position" -- ECR alone can't, it's position-relative by
#    definition. See the module docstring's DATA PROVENANCE note for the
#    honest caveat: PEAK/ALPHA are realistic curve-shape estimates (not
#    official 2026 projections), and REPLACEMENT_RANK is an explicit,
#    documented assumption about this league's real starter counts.
# =============================================================================

# Estimated full-season point ceiling for the #1 player at each position,
# under THIS league's exact scoring (0.5 PPR, 1.0 TE premium, Superflex
# demand baked into where we set QB's peak/replacement rather than a bolt-on
# multiplier). TE's peak is pushed close to WR/RB specifically because 1.0
# TE premium turns an elite pass-catching TE into a near-WR1 scorer.
PEAK_POINTS = {"QB": 420, "RB": 330, "WR": 310, "TE": 280}

# Power-law decay rate per position -- higher = faster value drop-off after
# the top of the position. RB and TE are historically the "cliff" positions
# (bell-cow/true-TE1 roles are scarce and irreplaceable); QB and WR stay
# comparatively deep, matching real draft-capital patterns.
DECAY_ALPHA = {"QB": 0.35, "RB": 0.55, "WR": 0.40, "TE": 0.65}

# Replacement rank = the last realistically-startable player at that
# position in THIS 10-team league, given the roster: 1 QB / 2 RB / 2 WR /
# 1 TE / 2 FLEX (RB/WR/TE) / 1 SUPERFLEX (QB/RB/WR/TE). These are explicit,
# documented assumptions about how those FLEX/SUPERFLEX slots actually get
# filled in practice -- edit them if your league behaves differently:
#   QB: 10 locked starters + ~9 of the 10 Superflex slots go to a QB2
#       (Superflex demand runs that high)                        -> ~19-20
#   RB: 20 locked starters + ~11 of the 20 FLEX slots + a little   -> ~30
#       Superflex overflow
#   WR: 20 locked starters + ~9 of the 20 FLEX slots               -> ~33-34
#   TE: 10 locked starters + ~3 of the 20 FLEX slots (TE premium    -> ~14-15
#       makes a 2nd TE flex-worthy more often than standard scoring)
REPLACEMENT_RANK = {"QB": 20, "RB": 30, "WR": 34, "TE": 15}

def _est_points(pos, ecr_pos):
    return PEAK_POINTS[pos] * (ecr_pos ** -DECAY_ALPHA[pos])

REPLACEMENT_POINTS = {pos: _est_points(pos, REPLACEMENT_RANK[pos]) for pos in PEAK_POINTS}

WEIGHTS = {
    "CY_BONUS": 0.05,       # +5% for contract-year motivation/usage
    "SOS_BONUS_MAX": 0.05,  # max +/-5% for a Wk1-4 schedule, scaled continuously by SoS_Rank (see below)
}

# ---- v7 (2026-08-29 night-before-draft safe upgrade #1): continuous SoS modifier.
# The old version applied a flat +/-5% based on which discrete 1-5 SoS_Tier a
# player landed in -- so a team ranked 6th-easiest and a team ranked 13th-easiest
# got the identical bonus, and a team ranked 7th got zero. SoS_Rank (1=easiest
# Wk1-4 slate .. 32=hardest) is continuous data we already compute; this just
# stops throwing that precision away. Linear ramp centered on rank 16.5 (the
# true midpoint of 32 teams): the single easiest schedule gets the full
# +SOS_BONUS_MAX, the single hardest gets the full -SOS_BONUS_MAX, and it
# scales smoothly in between instead of jumping in five flat steps.
def _sos_adj_continuous(row):
    if row["Team"] == "FA":
        return 1.0
    return 1 + WEIGHTS["SOS_BONUS_MAX"] * (16.5 - row["SoS_Rank"]) / 15.5

# ---- v7 safe upgrade #2: a narrow, HAND-CURATED durability discount.
# This is deliberately NOT a full-roster feature -- we do not have real
# games-missed data for all 328 players, and fabricating a durability score
# for players we haven't actually researched would be worse than having no
# discount at all (invisible fake precision). This dict only covers players
# where a real, cited injury/durability concern is already sitting in their
# ENRICHMENT Notes above (multi-year missed games, a recent ACL/Achilles
# tear, an active in-camp injury, a documented concussion history, etc.).
# Everyone else gets a neutral 1.0 -- no data, no discount, no guess.
DURABILITY = {
    "patrick mahomes": (0.97, "ACL scare"),
    "christian mccaffrey": (0.90, "repeated heavy-touch injury pattern at age 30"),
    "jahmyr gibbs": (0.98, "minor Aug hamstring issue, team says on schedule"),
    "jonathon brooks": (0.85, "2nd career ACL tear, minimal NFL touch history"),
    "zach charbonnet": (0.92, "recovering from an ACL tear"),
    "breece hall": (0.96, "Aug groin strain, expected back for Week 1"),
    "puka nacua": (0.93, "active psoas injury since mid-Aug"),
    "tee higgins": (0.90, "missed 12 games over 3 yrs incl. 2 concussions in 2025"),
    "malik nabers": (0.93, "recovering from a 2025 ACL tear"),
    "davante adams": (0.92, "age 34, snaps already shrinking pre-injury"),
    "chris bell": (0.93, "post-ACL"),
    "jordyn tyson": (0.85, "injured in all 4 college seasons plus a May injury"),
    "alec pierce": (0.85, "ankle injury complications, no clear return timeline"),
    "tyreek hill": (0.75, "severe 2025 multi-ligament knee injury, unsigned FA"),
    "chris olave": (0.92, "well-documented multi-concussion history"),
    "george kittle": (0.90, "torn Achilles in the 2025 season finale"),
}

def _durability_adj(row):
    return DURABILITY.get(normalize_name(row["Name"]), (1.0, ""))[0]

def _global_value(row):
    vbd = _est_points(row["Pos"], row["ECR_Pos"]) - REPLACEMENT_POINTS[row["Pos"]]
    cy_adj = 1 + (WEIGHTS["CY_BONUS"] if row["ContractYear"] == "Y" else 0)
    sos_adj = _sos_adj_continuous(row)
    dur_adj = _durability_adj(row)
    return round(vbd * cy_adj * sos_adj * dur_adj, 1)

df["GlobalValue"] = df.apply(_global_value, axis=1)
df["DurabilityNote"] = df["Name"].apply(lambda n: DURABILITY.get(normalize_name(n), (1.0, ""))[1])

# ---- 7a. GLOBAL TIER: one gap-analysis pass across the WHOLE 328-player
#      pool, sorted by GlobalValue descending -- this IS the board's row
#      order and primary Tier column.
df, GLOBAL_CLIFF_TIER = compute_tiers_generic(df, "GlobalValue", ascending=False, n_tiers=N_GLOBAL_TIERS)
df = df.rename(columns={"_tier_tmp": "Tier"}).reset_index(drop=True)
df.insert(0, "GlobalRank", range(1, len(df) + 1))

# ---- 7b. Persistent text Tier Level column (survives any manual sort/filter
#      in Excel, unlike the old row-position-based tier grouping) ----
df["TierLevel"] = df["Tier"].apply(lambda t: f"Tier {t}")

# ---- 7c. ADP + Value Delta -- real 2026 Superflex Half-PPR ADP (see
#      section 3b for provenance/caveats) matched in by normalized name.
#      Value Delta = ADP - GlobalRank: POSITIVE means the market is letting
#      this player fall PAST where our model has him (a real draft-day
#      slide -- great value); NEGATIVE means the market is taking him
#      EARLIER than our model would (a likely reach if you're the one
#      taking him there). Blank/NaN = no published ADP for that player
#      (expected for the deepest ~78 players in the pool).
ADP_MAP = {normalize_name(n): adp for n, team, adp in ADP_DATA}

df["ADP"] = df["Name"].apply(lambda n: ADP_MAP.get(normalize_name(n)))
df["ValueDelta"] = df.apply(
    lambda r: round(r["ADP"] - r["GlobalRank"], 1) if pd.notna(r["ADP"]) else None, axis=1)

df.to_csv(CSV_PATH, index=False)
print(f"Wrote {len(df)} players to {CSV_PATH}")
print(f"  -- {df['ADP'].notna().sum()} of {len(df)} players matched to real ADP data")

# =============================================================================
# 8. DRAFT PIVOT & ACTION RULES -- now generated PER GLOBAL TIER from the
#    actual position mix found in that tier (not a fixed per-position
#    script), so the guidance reacts to the real player pool: how many
#    Superflex-caliber QBs and elite TE-premium TEs are in THIS tier, and
#    how many of each remain in every tier after it.
# =============================================================================

# "Superflex-caliber" QB / "elite TE-premium" TE = PosTier <= these cutoffs
# (same thresholds the SF/TEP tag column already uses).
SF_QB_POS_TIER_CUTOFF = 4
ELITE_TE_POS_TIER_CUTOFF = 2

sf_qb_mask = (df["Pos"] == "QB") & (df["PosTier"] <= SF_QB_POS_TIER_CUTOFF)
elite_te_mask = (df["Pos"] == "TE") & (df["PosTier"] <= ELITE_TE_POS_TIER_CUTOFF)
TOTAL_SF_QB = int(sf_qb_mask.sum())
TOTAL_ELITE_TE = int(elite_te_mask.sum())
df["_sf_qb_cum"] = sf_qb_mask.cumsum()      # running count through this row, in GlobalRank order
df["_elite_te_cum"] = elite_te_mask.cumsum()

def _tier_qualifier(tier_num, n_tiers):
    frac = (tier_num - 1) / (n_tiers - 1) if n_tiers > 1 else 0
    if frac <= 0.15: return "Elite / Ceiling"
    if frac <= 0.35: return "Strong Starter / Floor"
    if frac <= 0.55: return "Startable / Boom-or-Bust"
    if frac <= 0.75: return "Value & Streamers"
    if frac <= 0.90: return "Bench Depth & Handcuffs"
    return "Deep Bench / Dart Throws"

def build_global_pivot_text(tier_num, tier_df, sf_qb_remaining_after, elite_te_remaining_after, is_cliff):
    sf_qb_in_tier = int(((tier_df["Pos"] == "QB") & (tier_df["PosTier"] <= SF_QB_POS_TIER_CUTOFF)).sum())
    elite_te_in_tier = int(((tier_df["Pos"] == "TE") & (tier_df["PosTier"] <= ELITE_TE_POS_TIER_CUTOFF)).sum())
    parts = []
    if sf_qb_in_tier > 0:
        parts.append(
            f"This tier has {sf_qb_in_tier} Superflex-caliber QB{'s' if sf_qb_in_tier != 1 else ''} in it -- "
            f"with 10 teams needing 2 startable arms each, take the QB here unless a clearly better "
            f"cross-positional GLOBAL VALUE is sitting in this same tier. Only {sf_qb_remaining_after} "
            f"Superflex-caliber QB{'s' if sf_qb_remaining_after != 1 else ''} remain in every tier after this one."
        )
    else:
        urgency = " -- getting thin, don't wait much longer" if sf_qb_remaining_after <= 6 else ""
        parts.append(
            f"No QB in this tier. {sf_qb_remaining_after} Superflex-caliber QB{'s' if sf_qb_remaining_after != 1 else ''} "
            f"remain in later tiers{urgency}. Don't reach for one here if the RB/WR/TE value in this tier "
            f"is clearly better -- but don't wait so long you're forced into true streamer territory at "
            f"your Superflex slot either."
        )
    if elite_te_in_tier > 0:
        parts.append(
            f"An elite TE-premium option is in this tier -- with 1.0 TE premium scoring this is one of "
            f"the very few TEs worth a real pick; passing means waiting on the cliff to true streamer TEs."
        )
    elif elite_te_remaining_after == 0:
        parts.append("Every true 1.0-TE-premium cheat-code TE is gone as of this tier -- punt TE to "
                      "streamer territory and reinforce QB/RB/WR instead.")
    if is_cliff:
        parts.append("⚠ CLIFF -- this is the single steepest value drop on the ENTIRE board. Grab someone "
                      "from this tier before a run at any position pushes you past it.")
    return " ".join(parts)

TIER_HEADER_PALETTE = ["2C6B4A", "33607F", "9A6B12", "B85A1E", "6A4A80", "555555",
                       "1F6B6B", "5A4A2C", "2E5A8A", "7A3B5A", "4A6B2C", "3A3A6B"]
FLAG_COLORS = {"STUD": "C6E8C6", "FLOOR": "C9DCEA", "SLEEPER": "F5E3AE", "GAMBLE": "F6D3B0",
               "BUST": "F1C2BC", "HANDCUFF": "E2D3EF", "DEPTH": "EAEAEA", "": "FFFFFF"}
SOS_COLORS = {1: "C6E8C6", 2: "DCEAD3", 3: "F2EFE3", 4: "F6D9C4", 5: "F1C2BC"}
CY_COLOR = "FFE58A"

# =============================================================================
# 9. BUILD THE WORKBOOK
# =============================================================================

wb = Workbook()
ws = wb.active
ws.title = "Draft Board"

HEADERS = ["DRAFTED?", "GLOBAL RANK", "TIER", "TIER LEVEL", "POS", "POS ECR", "POS TIER", "PLAYER",
           "TEAM", "BYE", "CY", "EARLY SoS", "SF / TEP", "ADP", "VALUE DELTA", "GLOBAL VALUE", "FLAG", "NOTES"]
N_COLS = len(HEADERS)
last_col_letter = get_column_letter(N_COLS)

col_widths = {"A": 11, "B": 8, "C": 6, "D": 9, "E": 6, "F": 8, "G": 8, "H": 22, "I": 7,
              "J": 6, "K": 6, "L": 11, "M": 22, "N": 8, "O": 11, "P": 11, "Q": 11, "R": 58}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# =============================================================================
# NOTE BANNER (row 1 only) -- the TIER GUIDE that used to live here (tier
# composition / cliffs / live remaining counts / Draft Pivot & Action Rule)
# moved to the top of the Instructions tab in v9, per explicit preference to
# keep the workbook at exactly 4 tabs. This sheet is now purely the clean,
# fully sortable/filterable player table: one pointer row, a blank spacer,
# then the header row, then data -- nothing else.
# =============================================================================

note_cell = ws.cell(row=1, column=1,
    value="TIER GUIDE (tier composition, cliffs, live remaining counts, and Draft Pivot & Action Rule "
          "per Global Tier) has moved to the Instructions tab -- open that tab for the always-visible "
          "reference. This sheet is just the player table below: a real, fully sortable/filterable "
          "Excel Table -- click any column's filter arrow to sort or filter by ADP, POS, PLAYER, "
          "GLOBAL VALUE, anything -- and the TIER LEVEL column keeps every player's tier visible "
          "regardless of how you've sorted.")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
note_cell.font = Font(bold=True, size=10.5, color="1C221B")
note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 40
# row 2 is left blank as a spacer between the note banner and the data table

# =============================================================================
# DATA TABLE -- a real, native Excel Table: one header row, uniform data rows,
# no merged cells in the body, so Excel's own sort/filter/autofilter dropdowns
# all just work (this is what "fully filterable" actually requires -- the old
# merged tier-banner rows interspersed among players could NOT be natively
# sorted/filtered without breaking).
# =============================================================================

table_header_row = 3
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=table_header_row, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="1C221B")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[table_header_row].height = 28

first_data_row = table_header_row + 1
r = first_data_row
for _, p in df.sort_values("GlobalRank").iterrows():
    values = ["", p["GlobalRank"], p["Tier"], p["TierLevel"], p["Pos"], p["ECR_Pos"], p["PosTier"],
              p["Name"], p["Team"], (int(p["Bye"]) if pd.notna(p["Bye"]) else "FA"),
              "YES" if p["ContractYear"] == "Y" else "", p["SoS_Label"],
              p["SF_TEP_Tag"], (p["ADP"] if pd.notna(p["ADP"]) else None),
              (p["ValueDelta"] if pd.notna(p["ValueDelta"]) else None),
              p["GlobalValue"], p["Flag"], p["Notes"]]
    for col, val in enumerate(values, start=1):
        ws.cell(row=r, column=col, value=val)
    for col in range(1, N_COLS + 1):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).font = Font(size=10)
    for col in (1, 2, 3, 4, 6, 7, 10, 14, 15, 16):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=18).alignment = Alignment(wrap_text=True, vertical="center")
    if p["ContractYear"] == "Y":
        ws.cell(row=r, column=11).fill = PatternFill("solid", fgColor=CY_COLOR)
        ws.cell(row=r, column=11).font = Font(size=10, bold=True)
    ws.cell(row=r, column=12).fill = PatternFill("solid", fgColor=SOS_COLORS[p["SoS_Tier"]])
    if p["Flag"]:
        ws.cell(row=r, column=17).fill = PatternFill("solid", fgColor=FLAG_COLORS[p["Flag"]])
        ws.cell(row=r, column=17).font = Font(size=10, bold=True)
    if p["SF_TEP_Tag"]:
        ws.cell(row=r, column=13).font = Font(size=9, italic=True)
    if pd.notna(p["ValueDelta"]):
        vd = p["ValueDelta"]
        vd_color = None
        if vd >= 15: vd_color = "8FD19E"     # big slide -- great value
        elif vd >= 5: vd_color = "D7ECDA"
        elif vd <= -15: vd_color = "F1A6A0"  # big reach risk
        elif vd <= -5: vd_color = "F7D7D3"
        if vd_color:
            ws.cell(row=r, column=15).fill = PatternFill("solid", fgColor=vd_color)
            ws.cell(row=r, column=15).font = Font(size=10, bold=True)
    r += 1
last_data_row = r - 1

dv = DataValidation(type="list", formula1='"✔"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"A{first_data_row}:A{last_data_row}")

strike_font = Font(strike=True, color="9A9A9A")
gray_fill = PatternFill("solid", fgColor="ECECEC")
ws.conditional_formatting.add(
    f"A{first_data_row}:{last_col_letter}{last_data_row}",
    FormulaRule(formula=[f'$A{first_data_row}="✔"'], font=strike_font, fill=gray_fill),
)

draft_table = Table(displayName="DraftBoardTable",
                     ref=f"A{table_header_row}:{last_col_letter}{last_data_row}")
draft_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                             showFirstColumn=False, showLastColumn=False,
                                             showColumnStripes=False)
ws.add_table(draft_table)

ws.freeze_panes = f"A{first_data_row}"

# ---------- Sheet 2: Scoring Rules -------------------------------------------
ws2 = wb.create_sheet("Scoring Rules")
ws2.append(["LEAGUE SETUP"]); ws2["A1"].font = Font(bold=True, size=14); ws2.append([])
for label, val in LEAGUE_INFO:
    ws2.append([label, val]); ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
ws2.append([]); ws2.append(["CUSTOM SCORING RULES"]); ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14)
ws2.append(["Category", "Rule", "Value"])
for c in ws2[ws2.max_row]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1C221B")
for cat, rule, val in SCORING_RULES:
    ws2.append([cat, rule, val])
cat_fill = {"PASSING": "E3ECF1", "RUSHING": "E1EEDD", "RECEIVING": "F3E7CC", "MISC": "EAEAEA"}
for row in ws2.iter_rows(min_row=ws2.max_row - len(SCORING_RULES) + 1, max_row=ws2.max_row):
    row[0].fill = PatternFill("solid", fgColor=cat_fill.get(row[0].value, "FFFFFF"))

# ---- Global Value (VBD) model, folded onto the bottom of this same sheet ---
ws2.append([]); ws2.append([])
ws2.append(["HOW THE DRAFT BOARD'S GLOBAL TIER / GLOBAL VALUE COLUMNS ARE CALCULATED"])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14)
ws2.append(["This is a Value-Based Drafting (VBD) model -- it's what makes QB/RB/WR/TE comparable on "
            "ONE scale so the whole board can be a single global order instead of four position lists."])
ws2.cell(row=ws2.max_row, column=1).font = Font(italic=True)
ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(wrap_text=True)
ws2.append(["Step 1: Est. Points = Peak Points x (Positional ECR Rank ^ -Decay Alpha)"])
ws2.append(["Step 2: Global Value = (Est. Points - Replacement Points for that position) x (1 + CY bonus) x "
            "(1 +/- continuous SoS bonus) x (Durability adj., default 1.0)"])
ws2.append([])
ws2.append(["Constant", "Value", "Meaning"])
for c in ws2[ws2.max_row]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1C221B")
value_model_rows = [
    ("PEAK_POINTS (QB)", PEAK_POINTS["QB"], "Estimated full-season points for the #1 QB -- Superflex demand is baked in here, not a bolt-on multiplier"),
    ("PEAK_POINTS (RB)", PEAK_POINTS["RB"], "Estimated full-season points for the #1 RB"),
    ("PEAK_POINTS (WR)", PEAK_POINTS["WR"], "Estimated full-season points for the #1 WR"),
    ("PEAK_POINTS (TE)", PEAK_POINTS["TE"], "Estimated full-season points for the #1 TE -- pushed close to WR/RB because 1.0 TE premium turns an elite pass-catcher into a near-WR1 scorer"),
    ("DECAY_ALPHA (QB)", DECAY_ALPHA["QB"], "How fast QB value falls off after #1 -- QB stays deep (Superflex demand)"),
    ("DECAY_ALPHA (RB)", DECAY_ALPHA["RB"], "How fast RB value falls off after #1 -- steepest cliff (bell-cow roles are scarce)"),
    ("DECAY_ALPHA (WR)", DECAY_ALPHA["WR"], "How fast WR value falls off after #1 -- stays deep, this format's safest floor position"),
    ("DECAY_ALPHA (TE)", DECAY_ALPHA["TE"], "How fast TE value falls off after #1 -- steep, true TE1 roles are scarce even with TE premium"),
    ("REPLACEMENT_RANK (QB)", REPLACEMENT_RANK["QB"], "Last realistically-startable QB in this league (~9 of 10 Superflex slots go to a QB2)"),
    ("REPLACEMENT_RANK (RB)", REPLACEMENT_RANK["RB"], "Last realistically-startable RB (2 locked + most of the FLEX allocation)"),
    ("REPLACEMENT_RANK (WR)", REPLACEMENT_RANK["WR"], "Last realistically-startable WR (2 locked + the rest of the FLEX allocation)"),
    ("REPLACEMENT_RANK (TE)", REPLACEMENT_RANK["TE"], "Last realistically-startable TE (1 locked + occasional FLEX use, elevated by TE premium)"),
    ("CY_BONUS", WEIGHTS["CY_BONUS"], "+/- applied when Contract Year = YES"),
    ("SOS_BONUS_MAX", WEIGHTS["SOS_BONUS_MAX"], "v7: CONTINUOUS now, not a step function -- the single easiest Wk1-4 slate "
     "(SoS_Rank 1) gets the full +max, the single hardest (rank 32) gets the full -max, scaled linearly in between"),
    ("DURABILITY", "0.75 - 1.00", "v7: a narrow, hand-curated discount (default 1.0, no discount) applied ONLY to the ~16 "
     "players whose Notes already cite a real, specific injury/durability concern (recent ACL/Achilles tear, multi-year "
     "missed games, an active in-camp injury, documented concussion history) -- not a full-roster feature, deliberately, "
     "since we don't have real games-missed data for all 328 players and won't fabricate it. See DurabilityNote in the CSV."),
]
for name, val, meaning in value_model_rows:
    ws2.append([name, val, meaning])
ws2.append([])
ws2.append(["These are plain constants near the top of build_draft_kit_2026.py (PEAK_POINTS, DECAY_ALPHA, "
            "REPLACEMENT_RANK, WEIGHTS) -- they are realistic, documented ESTIMATES of positional value "
            "shape and this league's real starter counts, not official 2026 point projections. Edit and "
            "re-run the script if your own read on positional scarcity differs -- e.g. if you think TE "
            "premium should close the gap to WR even further, raise PEAK_POINTS['TE']. The Draft Board's "
            "GLOBAL VALUE column is a static value, not a live formula, so it won't recalculate here."])
ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(wrap_text=True)
ws2.column_dimensions["A"].width = 24; ws2.column_dimensions["B"].width = 14; ws2.column_dimensions["C"].width = 90

# ---------- Sheet 3: Scoring Calculator --------------------------------------
ws3 = wb.create_sheet("Scoring Calculator")
ws3.append(["Paste a raw stat line below (edit the yellow cells) to see this league's exact fantasy-point total."])
ws3["A1"].font = Font(bold=True, size=12); ws3.append([])
calc_fields = [
    ("Is TE? (Y/N)", "N"),
    ("Pass Yds", 0), ("Pass TD (total)", 0), ("...of which 40+ yd TD", 0),
    ("INT thrown (total)", 0), ("...of which Pick-6 (returned for TD)", 0), ("Sacks Taken", 0),
    ("Rush Yds", 0), ("Rush TD (total)", 0), ("...of which 40+ yd TD", 0), ("2-PT Conversions", 0),
    ("Receptions", 0), ("Rec Yds", 0), ("Rec TD (total)", 0), ("...of which 40+ yd TD", 0),
    ("Fumbles (total)", 0), ("...of which Lost", 0),
]
input_row_start = 3
for i, (label, default) in enumerate(calc_fields):
    r = input_row_start + i
    ws3.cell(row=r, column=1, value=label).font = Font(bold=True)
    cell = ws3.cell(row=r, column=2, value=default)
    cell.fill = PatternFill("solid", fgColor="FFF6D9")
    cell.border = Border(*(Side(style="thin"),) * 4)

(R_TE, R_PASSYD, R_PASSTD, R_PASSTD40, R_INT, R_PICK6, R_SACK,
 R_RUSHYD, R_RUSHTD, R_RUSHTD40, R_2PT,
 R_REC, R_RECYD, R_RECTD, R_RECTD40,
 R_FUM, R_FUMLOST) = [input_row_start + i for i in range(len(calc_fields))]

formula = (
    f'=(B{R_PASSYD}*0.04) + (B{R_PASSTD}*5) + (B{R_PASSTD40}*0.5)'
    f' - (B{R_INT}*2) - (B{R_PICK6}*1) - (B{R_SACK}*0.5)'
    f' + (B{R_RUSHYD}*0.1) + (B{R_RUSHTD}*6) + (B{R_RUSHTD40}*2) + (B{R_2PT}*2)'
    f' + IF(B{R_RUSHYD}>=200, 4, IF(B{R_RUSHYD}>=100, 2, 0))'
    f' + (B{R_REC}*IF(UPPER(B{R_TE})="Y", 1, 0.5))'
    f' + (B{R_RECYD}*0.1) + (B{R_RECTD}*6) + (B{R_RECTD40}*2)'
    f' + IF(B{R_RECYD}>=200, 5, IF(B{R_RECYD}>=100, 3, 0))'
    f' - (B{R_FUM}*1) - (B{R_FUMLOST}*1)'
)
result_row = input_row_start + len(calc_fields) + 1
ws3.cell(row=result_row, column=1, value="TOTAL FANTASY POINTS").font = Font(bold=True, size=12)
result_cell = ws3.cell(row=result_row, column=2, value=formula)
result_cell.font = Font(bold=True, size=14, color="2C6B4A")
result_cell.fill = PatternFill("solid", fgColor="E4EEE3")
result_cell.number_format = "0.00"
ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 14

# ---------- Sheet 4: Instructions --------------------------------------------
ws4 = wb.create_sheet("Instructions")

for col, width in col_widths.items():
    ws4.column_dimensions[col].width = width

# =============================================================================
# TIER GUIDE (moved here from the Draft Board sheet in v9, per explicit
# preference to keep the workbook at exactly 4 tabs) -- a compact,
# ALWAYS-VISIBLE reference block: tier composition, cliff warnings, live
# remaining counts, and a data-driven Draft Pivot & Action Rule per tier.
# The live counts use cross-sheet COUNTIFS formulas back to the 'Draft
# Board' sheet's data table, since the guide and the table it counts now
# live on different tabs.
# =============================================================================

section_header_cell = ws4.cell(row=1, column=1, value="TIER GUIDE")
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
section_header_cell.font = Font(bold=True, size=14)
section_header_cell.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 20

title_cell = ws4.cell(row=2, column=1,
    value="Live counts + Draft Pivot & Action Rule per Global Tier (updates as you check off players "
          "on the Draft Board tab). The Draft Board tab has the actual player table -- a real, fully "
          "sortable/filterable Excel Table -- sort or filter by any column and the TIER LEVEL column "
          "there keeps every player's tier visible regardless.")
ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
title_cell.font = Font(bold=True, size=10.5, color="1C221B")
title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.row_dimensions[2].height = 40

guide_header_row = 4
ws4.merge_cells(start_row=guide_header_row, start_column=2, end_row=guide_header_row, end_column=5)
ws4.merge_cells(start_row=guide_header_row, start_column=7, end_row=guide_header_row, end_column=8)
ws4.merge_cells(start_row=guide_header_row, start_column=9, end_row=guide_header_row, end_column=N_COLS)
ws4.cell(row=guide_header_row, column=1, value="TIER")
ws4.cell(row=guide_header_row, column=2, value="COMPOSITION")
ws4.cell(row=guide_header_row, column=6, value="CLIFF")
ws4.cell(row=guide_header_row, column=7, value="REMAINING")
ws4.cell(row=guide_header_row, column=9, value="DRAFT PIVOT & ACTION RULE")
for c in range(1, N_COLS + 1):
    cell = ws4.cell(row=guide_header_row, column=c)
    if cell.value is None:
        continue
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor="1C221B")
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_idx = guide_header_row + 1
for tier in sorted(df["Tier"].unique()):
    tier_df = df[df["Tier"] == tier].sort_values("GlobalRank")
    is_cliff = (tier == GLOBAL_CLIFF_TIER)
    sf_qb_remaining_after = TOTAL_SF_QB - int(tier_df["_sf_qb_cum"].max())
    elite_te_remaining_after = TOTAL_ELITE_TE - int(tier_df["_elite_te_cum"].max())

    counts = tier_df["Pos"].value_counts().to_dict()
    comp_str = " / ".join(f"{counts[p]} {p}" for p in ["QB", "RB", "WR", "TE"] if p in counts)
    qualifier = _tier_qualifier(tier, N_GLOBAL_TIERS)
    pivot_text = build_global_pivot_text(tier, tier_df, sf_qb_remaining_after, elite_te_remaining_after, is_cliff)
    header_color = "9C3A2E" if is_cliff else TIER_HEADER_PALETTE[(tier - 1) % len(TIER_HEADER_PALETTE)]

    tier_cell = ws4.cell(row=row_idx, column=1, value=f"Tier {tier}")
    tier_cell.font = Font(bold=True, color="FFFFFF", size=10)
    tier_cell.fill = PatternFill("solid", fgColor=header_color)
    tier_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws4.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
    comp_cell = ws4.cell(row=row_idx, column=2, value=f"{qualifier}  ({comp_str})")
    comp_cell.font = Font(color="FFFFFF", size=9)
    comp_cell.fill = PatternFill("solid", fgColor=header_color)
    comp_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    cliff_cell = ws4.cell(row=row_idx, column=6, value="⚠ CLIFF ⚠" if is_cliff else "")
    cliff_cell.font = Font(bold=True, color="FFFFFF", size=9)
    cliff_cell.fill = PatternFill("solid", fgColor=header_color)
    cliff_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws4.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
    count_cell = ws4.cell(row=row_idx, column=7)
    count_cell.font = Font(bold=True, color="FFFFFF", size=9)
    count_cell.fill = PatternFill("solid", fgColor=header_color)
    count_cell.alignment = Alignment(horizontal="center", vertical="center")
    # cross-sheet COUNTIFS: the data table this now counts lives on 'Draft Board',
    # not on this sheet, since the Tier Guide moved here in v9.
    n_players = len(tier_df)
    count_cell.value = (
        f'="{n_players} total -- "&COUNTIFS(\'Draft Board\'!C{first_data_row}:C{last_data_row},{tier},'
        f'\'Draft Board\'!A{first_data_row}:A{last_data_row},"<>✔")&" still on the board"'
    )

    ws4.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx, end_column=N_COLS)
    pivot_cell = ws4.cell(row=row_idx, column=9, value=pivot_text)
    pivot_cell.font = Font(italic=True, color="FFFFFF", size=8.5)
    pivot_cell.fill = PatternFill("solid", fgColor=header_color)
    pivot_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws4.row_dimensions[row_idx].height = 34

    row_idx += 1

ws4.append([])  # blank spacer row between the Tier Guide and the instructions text below

lines = [
    ("HOW TO USE THIS DRAFT BOARD", True, 14), ("", False, 11),
    ("DRAFT-MORNING WATCH LIST (added 8/30, based on a fresh research pass the night before the draft)", True, 12),
    ("- Confirm active/practicing status before you draft: Jahmyr Gibbs (hamstring, pulled from practice after signing his "
     "extension -- 'on schedule' for the opener per the team) and Breece Hall (groin strain, out 8/17 -- expected back for "
     "Week 1 but not yet fully re-confirmed). Both are almost certainly fine; just don't assume, check.", False, 11),
    ("- ADP/VALUE DELTA blind spot: FantasyPros' ADP tool has no TE-premium toggle, so it prices Bowers/McBride/Kraft/Goedert "
     "off a non-TEP market. A real 2026 expert TE-premium Superflex mock had Bowers going 1.09 and McBride 2.09 -- earlier "
     "than this sheet's ADP column implies. Don't assume an elite TE-premium TE will slide to you like the Value Delta "
     "column suggests; if he's still there at your natural turn, that's the signal to take him, not to wait.", False, 11),
    ("- Superflex QB pacing: real 2026 Superflex mock drafts show most rooms have BOTH starting Superflex QB slots filled by "
     "around round 5-6. Don't let the Tier Guide's 'you can wait on QB2' language lull you past that point -- if you're "
     "into rounds 7-8 with only one QB rostered, treat your next good QB as a near-must-take over a marginal RB/WR/TE.", False, 11),
    ("", False, 11),
    ("ROSTER-AWARE MARGINAL VALUE GUIDE (v7, added the night before the draft)", True, 12),
    ("GLOBAL VALUE is computed once, up front, for the whole player pool -- it does not know what's already on YOUR "
     "roster, and it can't recalculate live as you draft (that would need an actual running program watching your picks "
     "in real time, which we deliberately did not rush into building and testing for the first time hours before a real "
     "draft -- ask Claude for the full reasoning if you want it). So apply this discount yourself, by eye, before locking "
     "in a pick: mentally knock a player's GLOBAL VALUE down by roughly this much based on how many you already have at "
     "that position.", False, 11),
    ("- QB: no discount for your 1st or 2nd (you need both the starter slot and the Superflex slot) -- discount a 3rd QB "
     "to roughly half its GLOBAL VALUE (bye-week/emergency insurance only), a 4th to almost nothing.", False, 11),
    ("- RB and WR: no discount through your first 3 -- discount a 4th to roughly 75% and a 5th+ to roughly 50-55% "
     "(still real bench depth and FLEX/handcuff value in this format, just not a starter-caliber add anymore).", False, 11),
    ("- TE: no discount for your 1st -- a 2nd is still real value (TE premium plus FLEX-eligibility keeps a 2nd TE "
     "around 75%) -- but discount a 3rd down to roughly 25-45%. You can only ever start 1 TE and occupy 1 FLEX with a "
     "second, so a 3rd elite TE mostly sits on your bench -- this is exactly what the 3-TE build in the mock-draft "
     "review got dinged for. Convert that pick into a startable RB/WR instead.", False, 11),
    ("", False, 11),
    ("1. TWO TABS WORK TOGETHER. The TIER GUIDE at the very top of THIS tab (Instructions) is a "
     "compact, always-visible reference showing each Global Tier's composition, cliff warning, live "
     "remaining count, and Draft Pivot & Action Rule -- keep this tab open next to Draft Board while "
     "you draft. The 'Draft Board' tab itself is just a one-line pointer banner plus the real player "
     "list, a native Excel TABLE -- click any column's filter arrow to sort or filter by ADP, POS, "
     "PLAYER, GLOBAL VALUE, anything. They're on separate tabs (rather than stacked on one sheet) so "
     "Excel's sort/filter never has to work around merged banner rows mixed into the data -- and so "
     "checking a player off on Draft Board still live-updates the remaining counts here via a "
     "cross-sheet formula.", False, 11),
    ("2. Drafted column (col A): click the cell, use the dropdown, pick the checkmark when a "
     "player is off the board. The whole row auto-strikes-through and grays out. Clear the cell "
     "(Delete key) to undo. This still works no matter how you've sorted/filtered the table.", False, 11),
    ("3. TIER LEVEL column ('Tier 1', 'Tier 2', ...) rides on every player row specifically so "
     "a player's tier stays visible even after you sort the table by ADP, name, or anything else -- "
     "you never lose track of tier grouping. TIER (numeric) is the same value, kept for sorting.", False, 11),
    ("4. ADP column: real 2026 Superflex Half-PPR consensus Average Draft Position (FantasyPros "
     "Real-Time ADP). VALUE DELTA = ADP - GLOBAL RANK: a big GREEN positive number means the "
     "market is letting that player slide well past where this model has him -- a real draft-day "
     "discount worth jumping on. A big RED negative number flags a likely reach if you take him "
     "there. Blank in both columns = no published ADP for that player (normal for the deepest "
     "~78 players in the pool -- see DATA PROVENANCE).", False, 11),
    ("5. TIER (global) is the primary tier column -- it's cut on the same cross-positional GLOBAL "
     "VALUE scale as the row order, using the same data-driven gap-analysis approach as before "
     "(biggest natural drop-offs become tier walls), just run once across all 328 players instead "
     "of four times. The single biggest gap on the ENTIRE board is the RED 'CLIFF' tier in the "
     "Tier Guide above.", False, 11),
    ("6. POS ECR and POS TIER are reference columns relative ONLY to a player's own position -- "
     "useful for 'who's the better WR' questions, but NOT for cross-position draft-order decisions. "
     "Use TIER/GLOBAL VALUE (or ADP/VALUE DELTA) for 'who should I actually take next.'", False, 11),
    ("7. The Tier Guide's Draft Pivot & Action Rule is DATA-DRIVEN per tier -- generated from what's "
     "actually in that tier (how many Superflex-caliber QBs, whether an elite TE-premium TE is "
     "mixed in, how many of each remain in later tiers), and its live remaining-count updates "
     "automatically as you check off players on the Draft Board tab, regardless of sort order.", False, 11),
    ("8. CY column: gold highlight = final year of that player's deal (rookie or veteran).", False, 11),
    ("9. EARLY SoS: color-coded schedule difficulty (green=soft, red=gauntlet), isolated to "
     "WEEKS 1-4 ONLY -- built from each team's real 2026 Weeks 1-4 opponents, not a season-long "
     "blend. The 4 real opponents' pass/run defensive strength is averaged and re-ranked across "
     "all 32 teams. See the script's DATA PROVENANCE note for the honest caveat on the defense-"
     "strength inputs (real final-2025 yards-allowed proxy, no paid 2026 DVOA split exists free).", False, 11),
    ("10. SF / TEP column: flags Superflex-relevant QBs and TE-premium cheat-code TEs (based on "
     "POS TIER, since 'is this an elite QB/TE' is a within-position question), plus a bonus "
     "'SF Overall #n' tag for the ~20 players whose true Superflex-specific ECR rank was publicly "
     "available (FantasyPros paywalls the rest of that list).", False, 11),
    ("11. GLOBAL VALUE column: a Value-Based-Drafting (VBD) score -- ECR converted to an estimated "
     "points curve per position, minus a replacement-level baseline for THIS league's real starter "
     "counts -- see the bottom of the 'Scoring Rules' sheet for the exact formula and constants.", False, 11),
    ("12. FLAG column: STUD / FLOOR / SLEEPER / GAMBLE / BUST / HANDCUFF / DEPTH, carried over "
     "from the Aug 24-29, 2026 research pass. A player can be a high Tier (real talent/ECR) AND "
     "flagged BUST (i.e. being drafted at a price the current situation doesn't support).", False, 11),
    ("13. Use the filter arrows on the table's header row to show just one position, sort by ADP "
     "or VALUE DELTA, hide drafted players, etc. A note on your 8th-overall pick: with 7 picks "
     "ahead of you in round 1, expect the front of Global Tier 1 to already be gone by 1.08 -- "
     "treat the back of Tier 1 and the front of Tier 2 as your realistic target range there.", False, 11),
    ("14. BYE column: that player's real 2026 bye week (verified against 2 independent sources -- "
     "there is no Week 12 bye league-wide this season). \"FA\" means that player has no current NFL "
     "team, so there is genuinely no bye week to report (Tyreek Hill, Kareem Hunt, Joe Mixon, Cedric "
     "Tillman) -- that's a verified absence, not missing data. Purely informational -- it does not "
     "feed into GLOBAL VALUE or tiering. Use it to sanity-check your roster once you've got starters "
     "at each position, so you're not left scrambling for a fill-in when several of them share a bye "
     "-- Week 11 is this year's most crowded bye week (6 teams), worth watching for.", False, 11),
    ("", False, 11),
    ("EXTENDING OR CORRECTING THE PLAYER POOL", True, 13),
    ("The full 328-player ECR pool covers QB/RB/WR/TE deep into replacement level -- players.csv "
     "is the source of truth (GlobalRank, Name, Team, Pos, ECR_Pos, PosTier, TierLevel, Bye, "
     "ContractYear, Flag, Notes, SoS_*, ADP, ValueDelta, GlobalValue, Tier). To hand-edit a tier, "
     "flag, ADP, or note, change it there and re-run the script with that file wired in as the "
     "source (or just ask Claude to do it).", False, 11),
    ("", False, 11),
    ("DATA PROVENANCE", True, 13),
    ("ECR + team: FantasyPros Half-PPR Overall Draft ECR, 108 experts, fetched 2026-08-29. "
     "Situational tags were hand-researched the week of Aug 24-29, 2026 from RotoWire, "
     "FantasyPros, CBS Sports, Yahoo, PhillyVoice, NFL.com, and the league manager's own "
     "contract-year depth-chart screenshots -- only players from that pass carry a Flag/rich "
     "Note; everyone else in the ECR pool is unflagged but still correctly ECR-ranked. "
     "Re-verify anything time-sensitive the morning of your draft. GLOBAL VALUE is a VBD estimate "
     "built on realistic but hand-set positional point curves, not an official 2026 projection -- "
     "see the 'Scoring Rules' sheet for the full methodology and constants. ADP: FantasyPros "
     "Real-Time ADP, Superflex view, Half-PPR, fetched 2026-08-29, 12-team (no 10-team toggle was "
     "available -- treat it as a relative ranking, not a literal pick slot), capped at 250 real "
     "players so the deepest ~78 players in the pool have no ADP/Value Delta. DRAFT-MORNING REFRESH (8/30): the Flag/Notes "
     "for Bucky Irving, Matthew Golden, David Montgomery, Rhamondre Stevenson, Kenny Gainwell, Breece Hall, Chris Olave, "
     "Jonah Coleman, Jalen Coker, Tyler Allgeier, Dallas Goedert, Jahmyr Gibbs, Trevor Lawrence, Bryce Young, Brian Thomas "
     "Jr., Jakobi Meyers, Michael Pittman Jr., Rashid Shaheed, Jaylen Waddle, and Romeo Doubs were corrected or added from "
     "a same-night research pass (training camp reports, injury news, trades) -- look for 'NEW 8/30' / 'UPDATED 8/30' in "
     "their Notes.", False, 11),
]
for text, bold, size in lines:
    ws4.append([text])
    row = ws4.max_row
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    c = ws4.cell(row=row, column=1)
    c.font = Font(bold=bold, size=size)
    c.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(XLSX_PATH)
print(f"Wrote {XLSX_PATH}")
